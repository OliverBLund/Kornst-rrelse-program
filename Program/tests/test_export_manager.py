"""
Regression tests for export manager CSV filtering behavior.
"""

import csv
import json
import os
import sys
import tempfile
import unittest
from unittest.mock import Mock, patch

os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
sys.path.insert(0, 'Program')

from data_loader import GrainSizeData
from matplotlib.figure import Figure
from PyQt6.QtWidgets import QApplication, QPushButton, QTableWidget, QTabWidget
from gui.export_tab import ExportTab
from gui.export_manager import ExportManager
from gui.plot_styles import PROFESSIONAL_STYLE
from k_calculations_v2 import CalculationStatus, KCalculationResult
from unit_conversions import HydraulicConductivityConverter, HydraulicConductivityUnit

APP = QApplication.instance() or QApplication([])


def build_dataset(name: str = 'Sample A') -> GrainSizeData:
    """Create a stable grain-size dataset for export tests."""
    return GrainSizeData(
        sample_name=name,
        temperature=20.0,
        porosity=0.35,
        particle_sizes=[4.75, 2.0, 1.0, 0.5, 0.25, 0.125, 0.063],
        percent_passing=[100.0, 95.0, 82.0, 61.0, 38.0, 14.0, 4.0],
    )


def build_results() -> list[KCalculationResult]:
    """Create representative K-results spanning multiple method groups."""
    return [
        KCalculationResult(
            method_name='Hazen',
            k_value=1.0e-4,
            formula_used='k = c * d10^2',
            status=CalculationStatus.OK,
            status_message='',
            conditions_met=True,
            temperature=20.0,
            porosity=0.35,
            grain_size_used='D10',
        ),
        KCalculationResult(
            method_name='Beyer',
            k_value=1.5e-4,
            formula_used='k = c * d10^2 * log(500 / Cu)',
            status=CalculationStatus.OK,
            status_message='',
            conditions_met=True,
            temperature=20.0,
            porosity=0.35,
            grain_size_used='D10',
        ),
        KCalculationResult(
            method_name='USBR',
            k_value=2.5e-4,
            formula_used='k = c * d20^2.3',
            status=CalculationStatus.OK,
            status_message='',
            conditions_met=True,
            temperature=20.0,
            porosity=0.35,
            grain_size_used='D20',
        ),
    ]


def read_csv_rows(path: str) -> list[list[str]]:
    with open(path, newline='', encoding='utf-8') as handle:
        return list(csv.reader(handle))


def read_json(path: str) -> dict:
    with open(path, encoding='utf-8') as handle:
        return json.load(handle)


class RecordingProgress:
    def __init__(self):
        self.maximum = None
        self.values = []

    def setMaximum(self, value):
        self.maximum = value

    def setValue(self, value):
        self.values.append(value)


class TestExportManagerExports(unittest.TestCase):
    def setUp(self):
        self.dataset = build_dataset()
        self.results = build_results()
        self.datasets = [(self.dataset.sample_name, self.dataset, self.results)]

    def make_config(self, output_dir: str, **overrides):
        config = {
            'csv': True,
            'csv_mode': 'combined',
            'csv_long': True,
            'csv_wide': False,
            'excel': False,
            'json': False,
            'png': False,
            'svg': False,
            'pdf_plot': False,
            'grain_distribution': False,
            'percentiles': True,
            'gradation': True,
            'classification': False,
            'k_values': True,
            'statistics': False,
            'plots': False,
            'formulas': False,
            'validation': False,
            'k_filter_mode': 'all',
            'selected_k_categories': {
                'hazen_based': True,
                'porosity_dependent': True,
                'uniformity_dependent': True,
                'empirical': True,
                'temperature_corrected': True,
            },
            'selected_percentiles': ['d10', 'd20', 'd30', 'd50', 'd60'],
            'selected_k_methods': None,
            'k_units': {
                'm_s': True,
                'cm_s': True,
                'm_d': True,
            },
            'selected_statistics': ['mean', 'median', 'std_dev', 'min', 'max', 'valid_count'],
            'include_grain_size_stats': True,
            'include_metadata': {
                'sample_info': True,
                'environmental': True,
                'export_timestamp': True,
            },
            'output_dir': output_dir,
            'filename_template': '{sample_name}_results_{date}',
        }
        config.update(overrides)
        return config

    def test_combined_long_csv_respects_selected_methods_and_units(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv_long=True,
                csv_wide=False,
                k_filter_mode='individual',
                selected_k_methods=['Beyer', 'USBR'],
                k_units={'m_s': True, 'cm_s': False, 'm_d': True},
                selected_percentiles=['d10', 'd95'],
                gradation=False,
            )

            exported = ExportManager().export(self.datasets, config)

            self.assertEqual(len(exported), 1)
            export_name = os.path.basename(exported[0])
            self.assertIn('combined_all_datasets', export_name)

            rows = read_csv_rows(exported[0])
            header = rows[0]
            methods = [row[1] for row in rows[1:]]

            self.assertIn(os.path.join('tables', 'csv'), exported[0])
            self.assertIn('K (m/s)', header)
            self.assertIn('K (m/d)', header)
            self.assertNotIn('K (cm/s)', header)
            self.assertIn('D10 (mm)', header)
            self.assertIn('D95 (mm)', header)
            self.assertNotIn('D20 (mm)', header)
            self.assertNotIn('D50 (mm)', header)
            self.assertNotIn('Cu', header)
            self.assertNotIn('Cc', header)
            self.assertEqual(methods, ['Beyer', 'USBR'])
            self.assertTrue(all(len(row) == len(header) for row in rows[1:]))

    def test_wide_csv_only_writes_requested_format_and_category_methods(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv_long=False,
                csv_wide=True,
                statistics=True,
                classification=True,
                k_filter_mode='category',
                selected_k_categories={
                    'hazen_based': False,
                    'porosity_dependent': False,
                    'uniformity_dependent': True,
                    'empirical': False,
                    'temperature_corrected': False,
                },
                selected_percentiles=['d10', 'd60'],
                k_units={'m_s': False, 'cm_s': True, 'm_d': False},
                selected_statistics=['median', 'valid_count'],
            )

            exported = ExportManager().export(self.datasets, config)

            self.assertEqual(len(exported), 1)
            export_name = os.path.basename(exported[0])
            self.assertIn('wide_format_all_datasets', export_name)

            rows = read_csv_rows(exported[0])
            header = rows[0]
            data_row = rows[1]

            self.assertIn('D10_mm', header)
            self.assertIn('D60_mm', header)
            self.assertIn('Soil_Classification', header)
            self.assertNotIn('D20_mm', header)
            self.assertIn('K_Beyer_cm/s', header)
            self.assertIn('Status_Beyer', header)
            self.assertIn('K_Median_cm/s', header)
            self.assertNotIn('K_Mean_cm/s', header)
            self.assertNotIn('K_Hazen_cm/s', header)
            self.assertNotIn('K_Beyer_m/s', header)
            self.assertNotIn('K_Beyer_m/d', header)
            self.assertEqual(data_row[header.index('Valid_Methods_Count')], '1')

    def test_long_csv_omits_k_and_grain_columns_when_content_is_disabled(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv_long=True,
                csv_wide=False,
                k_values=False,
                statistics=False,
                percentiles=False,
                gradation=False,
                classification=False,
            )

            exported = ExportManager().export(self.datasets, config)

            rows = read_csv_rows(exported[0])
            header = rows[0]
            self.assertEqual(len(rows), 2)
            self.assertNotIn('Method', header)
            self.assertNotIn('Status', header)
            self.assertFalse(any(column.startswith('K ') for column in header))
            self.assertNotIn('D10 (mm)', header)
            self.assertNotIn('Cu', header)
            self.assertNotIn('Soil Classification', header)

    def test_wide_csv_omits_k_statistics_and_grain_columns_when_content_is_disabled(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv_long=False,
                csv_wide=True,
                k_values=False,
                statistics=False,
                percentiles=False,
                gradation=False,
                classification=False,
            )

            exported = ExportManager().export(self.datasets, config)

            rows = read_csv_rows(exported[0])
            header = rows[0]
            self.assertNotIn('K_Hazen_m/s', header)
            self.assertNotIn('Status_Hazen', header)
            self.assertNotIn('K_Mean_m/s', header)
            self.assertNotIn('Valid_Methods_Count', header)
            self.assertNotIn('D10_mm', header)
            self.assertNotIn('Cu_Uniformity_Coefficient', header)

    def test_separate_statistics_csv_uses_filtered_results_metadata_and_stat_selection(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv_mode='separate',
                csv_long=True,
                csv_wide=False,
                grain_distribution=False,
                k_values=False,
                statistics=True,
                k_filter_mode='individual',
                selected_k_methods=['USBR'],
                k_units={'m_s': False, 'cm_s': False, 'm_d': True},
                selected_statistics=['median', 'valid_count'],
                include_grain_size_stats=False,
                include_metadata={
                    'sample_info': False,
                    'environmental': False,
                    'export_timestamp': True,
                },
            )

            exported = ExportManager().export(self.datasets, config)

            self.assertEqual(len(exported), 1)
            export_name = os.path.basename(exported[0])
            self.assertIn('statistics', export_name)

            rows = read_csv_rows(exported[0])
            timestamp_rows = [row for row in rows if row and row[0] == 'Export Timestamp']
            self.assertEqual(len(timestamp_rows), 1)
            self.assertTrue(timestamp_rows[0][1])
            self.assertIn(['Statistic', 'K (m/d)'], rows)
            self.assertIn(['Median', '21.60'], rows)
            self.assertIn(['Valid Count', '1'], rows)
            self.assertNotIn(['Mean', '21.60'], rows)
            self.assertNotIn(['Sample Name', self.dataset.sample_name], rows)
            self.assertFalse(any(row and row[0] == 'Grain Size Percentiles' for row in rows))

    def test_json_export_honors_metadata_units_and_selected_statistics(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv=False,
                json=True,
                statistics=True,
                k_filter_mode='individual',
                selected_k_methods=['USBR'],
                selected_percentiles=['d10', 'd95'],
                k_units={'m_s': False, 'cm_s': True, 'm_d': False},
                selected_statistics=['median', 'valid_count'],
                include_metadata={
                    'sample_info': False,
                    'environmental': False,
                    'export_timestamp': True,
                },
            )

            exported = ExportManager().export(self.datasets, config)

            self.assertEqual(len(exported), 1)
            payload = read_json(exported[0])

            self.assertNotIn('sample_name', payload)
            self.assertEqual(set(payload['metadata'].keys()), {'exported_at'})
            self.assertTrue(payload['metadata']['exported_at'])
            self.assertEqual(len(payload['k_values']), 1)
            self.assertEqual(payload['k_values'][0]['method'], 'USBR')
            self.assertIn('k_cm_s', payload['k_values'][0])
            self.assertNotIn('k_m_s', payload['k_values'][0])
            self.assertNotIn('k_m_d', payload['k_values'][0])
            self.assertEqual(set(payload['percentiles'].keys()), {'D10', 'D95'})
            self.assertEqual(set(payload['statistics'].keys()), {'median_k_cm_s', 'valid_count'})

    def test_json_export_records_effective_porosity_and_source(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            dataset = build_dataset()
            dataset.calculated_porosity = 0.3123
            dataset.current_porosity = 0.5000
            dataset.porosity = 0.3500
            datasets = [(dataset.sample_name, dataset, self.results)]
            config = self.make_config(
                temp_dir,
                csv=False,
                json=True,
                statistics=False,
                include_metadata={
                    'sample_info': True,
                    'environmental': True,
                    'export_timestamp': False,
                },
            )

            exported = ExportManager().export(datasets, config)
            payload = read_json(exported[0])

            self.assertEqual(payload['metadata']['porosity'], 0.5)
            self.assertIn('Manual override', payload['metadata']['porosity_source'])


    def test_combined_csv_exports_raw_grain_distribution_table(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv_long=True,
                csv_wide=False,
                grain_distribution=True,
            )

            exported = ExportManager().export(self.datasets, config)

            grain_path = next(path for path in exported if 'grain_distribution' in os.path.basename(path))
            rows = read_csv_rows(grain_path)
            self.assertEqual(rows[0], ['Sample Name', 'Point', 'Particle Size (mm)', 'Percent Passing (%)'])
            self.assertEqual(rows[1], ['Sample A', '1', '4.75', '100.0'])
            self.assertEqual(len(rows), len(self.dataset.particle_sizes) + 1)

    def test_combined_excel_uses_shared_typed_tables(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest('openpyxl not installed')

        dataset_b = build_dataset('Sample B')
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv=False,
                excel=True,
                excel_mode='combined',
                grain_distribution=True,
                statistics=True,
            )

            exported = ExportManager().export(datasets, config)

            self.assertEqual(len(exported), 1)
            workbook = load_workbook(exported[0], data_only=True)
            self.assertIn('K_Results_Long', workbook.sheetnames)
            self.assertIn('Sample_Wide', workbook.sheetnames)
            self.assertIn('Grain_Distribution', workbook.sheetnames)
            self.assertIn('Aggregate_Statistics', workbook.sheetnames)

            wide_rows = list(workbook['Sample_Wide'].iter_rows(values_only=True))
            header = list(wide_rows[0])
            hazen_col = header.index('K_Hazen_m/s')
            self.assertIsInstance(wide_rows[1][hazen_col], float)
            self.assertAlmostEqual(wide_rows[1][hazen_col], 1.0e-4)

            grain_rows = list(workbook['Grain_Distribution'].iter_rows(values_only=True))
            self.assertIsInstance(grain_rows[1][2], float)
            self.assertIsInstance(grain_rows[1][3], (int, float))

    def test_duplicate_sample_names_get_stable_plot_file_labels(self):
        duplicate = build_dataset('Sample A')
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (duplicate.sample_name, duplicate, build_results()),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv=False,
                png=True,
                plots=True,
                selected_plot_types=['grain_size_curve'],
            )

            exported = ExportManager().export(datasets, config)

            self.assertEqual(len(exported), 2)
            self.assertTrue(any('Sample A (1)' in path for path in exported))
            self.assertTrue(any('Sample A (2)' in path for path in exported))

    def test_export_statistics_use_shared_ok_only_summary(self):
        warning_result = KCalculationResult(
            method_name='Kruger',
            k_value=1.0e-2,
            formula_used='',
            status=CalculationStatus.WARNING,
            status_message='outside recommended range',
            conditions_met=False,
            temperature=20.0,
            porosity=0.35,
            grain_size_used='D50',
        )
        ok_low = KCalculationResult(
            method_name='Hazen',
            k_value=1.0e-4,
            formula_used='',
            status=CalculationStatus.OK,
            status_message='',
            conditions_met=True,
            temperature=20.0,
            porosity=0.35,
            grain_size_used='D10',
        )
        ok_high = KCalculationResult(
            method_name='USBR',
            k_value=4.0e-4,
            formula_used='',
            status=CalculationStatus.OK,
            status_message='',
            conditions_met=True,
            temperature=20.0,
            porosity=0.35,
            grain_size_used='D20',
        )

        config = self.make_config(
            tempfile.gettempdir(),
            selected_statistics=['geometric_mean', 'mean', 'valid_count'],
        )
        stats = ExportManager()._calculate_statistics(
            [ok_low, warning_result, ok_high],
            config,
        )

        self.assertAlmostEqual(stats['geometric_mean_k'], 2.0e-4)
        self.assertAlmostEqual(stats['mean_k'], 2.5e-4)
        self.assertEqual(stats['valid_count'], 2)

    def test_aggregate_statistics_table_includes_overall_groups_and_datasets(self):
        dataset_b = build_dataset('Sample B')
        self.dataset.group_name = 'Layer A'
        dataset_b.group_name = 'Layer B'
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        config = self.make_config(tempfile.gettempdir(), selected_statistics=['geometric_mean', 'mean', 'valid_count'])

        rows = ExportManager().build_aggregate_statistics_table(datasets, config)

        headers = rows[0]
        self.assertIn('Scope_Type', headers)
        self.assertIn('K_Geometric_Mean_m_s', headers)
        self.assertIn('D50_Median_mm', headers)
        scope_pairs = {(row[0], row[1]) for row in rows[1:]}
        self.assertIn(('Overall', 'Overall'), scope_pairs)
        self.assertIn(('Group', 'Layer A'), scope_pairs)
        self.assertIn(('Group', 'Layer B'), scope_pairs)
        self.assertIn(('Dataset', 'Sample A'), scope_pairs)
        self.assertIn(('Dataset', 'Sample B'), scope_pairs)

    def test_csv_export_writes_collection_aggregate_statistics(self):
        dataset_b = build_dataset('Sample B')
        self.dataset.group_name = 'Layer A'
        dataset_b.group_name = 'Layer B'
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(temp_dir, csv_long=True, csv_wide=False, excel=False, statistics=True)

            exported = ExportManager().export(datasets, config)

            aggregate_path = next(path for path in exported if 'aggregate_statistics' in os.path.basename(path))
            rows = read_csv_rows(aggregate_path)
            scope_pairs = {(row[0], row[1]) for row in rows[1:]}
            self.assertIn(('Overall', 'Overall'), scope_pairs)
            self.assertIn(('Group', 'Layer A'), scope_pairs)
            self.assertIn(('Group', 'Layer B'), scope_pairs)

    def test_collection_aggregate_export_can_be_disabled(self):
        dataset_b = build_dataset('Sample B')
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv_long=True,
                csv_wide=False,
                excel=False,
                statistics=True,
                include_collection_aggregates=False,
            )

            exported = ExportManager().export(datasets, config)

            self.assertFalse(any('aggregate_statistics' in os.path.basename(path) for path in exported))

    def test_excel_export_writes_collection_aggregate_statistics_workbook(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest('openpyxl not installed')

        dataset_b = build_dataset('Sample B')
        self.dataset.group_name = 'Layer A'
        dataset_b.group_name = 'Layer B'
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(temp_dir, csv=False, excel=True, statistics=True)

            exported = ExportManager().export(datasets, config)

            aggregate_path = next(path for path in exported if 'aggregate_statistics' in os.path.basename(path))
            workbook = load_workbook(aggregate_path, data_only=True)
            rows = list(workbook['Aggregate_Statistics'].iter_rows(values_only=True))
            scope_pairs = {(row[0], row[1]) for row in rows[1:]}
            self.assertIn(('Overall', 'Overall'), scope_pairs)
            self.assertIn(('Group', 'Layer A'), scope_pairs)
            self.assertIn(('Dataset', 'Sample B'), scope_pairs)

    def test_excel_export_honors_metadata_and_stat_selection(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest('openpyxl not installed')

        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv=False,
                excel=True,
                statistics=True,
                k_units={'m_s': False, 'cm_s': False, 'm_d': True},
                selected_statistics=['median', 'valid_count'],
                include_grain_size_stats=False,
                include_metadata={
                    'sample_info': False,
                    'environmental': True,
                    'export_timestamp': True,
                },
            )

            exported = ExportManager().export(self.datasets, config)

            self.assertEqual(len(exported), 1)
            workbook = load_workbook(exported[0], data_only=True)
            summary_values = [row[0] for row in workbook['Summary'].iter_rows(values_only=True) if row and row[0]]
            stats_rows = list(workbook['Statistics'].iter_rows(values_only=True))

            self.assertNotIn('Sample Name:', summary_values)
            self.assertIn('Temperature:', summary_values)
            self.assertIn('Exported At:', summary_values)
            self.assertIn('Median (m/d):', summary_values)
            self.assertNotIn('Mean (m/d):', summary_values)
            self.assertEqual(stats_rows[0], ('Statistic', 'K (m/d)'))
            self.assertIn(('Median', 12.96), stats_rows)
            self.assertIn(('Valid Count', 3), stats_rows)
            self.assertFalse(any(row and row[0] == 'Mean' for row in stats_rows[1:]))

    def test_excel_export_honors_percentile_method_unit_and_detail_selection(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest('openpyxl not installed')

        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv=False,
                excel=True,
                statistics=False,
                selected_percentiles=['d10', 'd95'],
                k_filter_mode='individual',
                selected_k_methods=['USBR'],
                k_units={'m_s': False, 'cm_s': True, 'm_d': False},
                formulas=True,
                validation=True,
            )

            exported = ExportManager().export(self.datasets, config)

            self.assertEqual(len(exported), 1)
            workbook = load_workbook(exported[0], data_only=True)

            summary_values = [
                row[0]
                for row in workbook['Summary'].iter_rows(values_only=True)
                if row and row[0]
            ]
            percentile_rows = [
                row
                for row in workbook['Percentiles'].iter_rows(values_only=True)
                if row and row[0]
            ]
            k_rows = [
                row
                for row in workbook['K_Values'].iter_rows(values_only=True)
                if row and row[0]
            ]

            self.assertIn('D10:', summary_values)
            self.assertIn('D95:', summary_values)
            self.assertNotIn('D50:', summary_values)
            self.assertEqual(percentile_rows[0], ('Percentile', 'Size (mm)'))
            self.assertEqual([row[0] for row in percentile_rows[1:]], ['D10', 'D95'])
            self.assertEqual(
                k_rows[0],
                ('Method', 'K (cm/s)', 'Status', 'Status Message', 'Formula'),
            )
            self.assertEqual([row[0] for row in k_rows[1:]], ['USBR'])

    def test_plot_export_generates_png_without_live_figure(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv=False,
                png=True,
                plots=True,
                plot_figures=[],
                plot_contexts=[],
            )

            exported = ExportManager().export(self.datasets, config)

            self.assertEqual(len(exported), 1)
            self.assertTrue(exported[0].endswith('_plot.png'))
            self.assertIn(os.path.join('plots', self.dataset.sample_name), exported[0])
            with open(exported[0], 'rb') as handle:
                self.assertEqual(handle.read(8), b'\x89PNG\r\n\x1a\n')

    def test_plot_save_uses_white_report_background(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(temp_dir, csv=False, png=True)
            manager = ExportManager()
            figure = Figure()
            figure.add_subplot(111).plot([1, 2], [1, 2])
            figure.savefig = Mock()

            manager._save_plot_figure(figure, 'sample', 'grain_size_curve', config)

            self.assertEqual(figure.patch.get_facecolor(), (1.0, 1.0, 1.0, 1.0))
            _, kwargs = figure.savefig.call_args
            self.assertEqual(kwargs['facecolor'], 'white')
            self.assertEqual(kwargs['edgecolor'], 'white')
            self.assertIn(os.path.join('plots', 'collection'), figure.savefig.call_args[0][0])

    def test_plot_export_generates_vector_formats_without_live_figure(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv=False,
                svg=True,
                pdf_plot=True,
                plots=True,
                plot_figures=[],
                plot_contexts=[],
            )

            exported = ExportManager().export(self.datasets, config)

            self.assertEqual({os.path.splitext(path)[1] for path in exported}, {'.svg', '.pdf'})
            svg_path = next(path for path in exported if path.endswith('.svg'))
            pdf_path = next(path for path in exported if path.endswith('.pdf'))
            with open(svg_path, encoding='utf-8') as handle:
                self.assertIn('<svg', handle.read())
            with open(pdf_path, 'rb') as handle:
                self.assertEqual(handle.read(4), b'%PDF')

    def test_plot_export_writes_selected_single_sample_plot_types(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv=False,
                png=True,
                plots=True,
                selected_plot_types=[
                    'grain_size_curve',
                    'k_value_bar',
                    'applicability_heatmap',
                ],
            )

            exported = ExportManager().export(self.datasets, config)

            basenames = {os.path.basename(path) for path in exported}
            self.assertEqual(len(exported), 3)
            self.assertTrue(any(name.endswith('_plot.png') for name in basenames))
            self.assertTrue(any(name.endswith('_k_values.png') for name in basenames))
            self.assertTrue(any(name.endswith('_applicability.png') for name in basenames))

    def test_plot_export_writes_selected_collection_plot_types_once(self):
        dataset_b = build_dataset('Sample B')
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv=False,
                png=True,
                plots=True,
                selected_plot_types=[
                    'distribution_overlay',
                    'k_value_comparison',
                    'statistical_boxplots',
                    'reliability_matrix',
                ],
            )

            exported = ExportManager().export(datasets, config)

            basenames = {os.path.basename(path) for path in exported}
            self.assertEqual(len(exported), 4)
            self.assertTrue(any(name.endswith('_distribution_overlay.png') for name in basenames))
            self.assertTrue(any(name.endswith('_k_value_comparison.png') for name in basenames))
            self.assertTrue(any(name.endswith('_k_value_boxplot.png') for name in basenames))
            self.assertTrue(any(name.endswith('_reliability_matrix.png') for name in basenames))
            self.assertTrue(all(name.startswith('all_datasets_results_') for name in basenames))

    def test_k_distribution_collection_plot_renders_lognormal_histogram(self):
        dataset_b = build_dataset('Sample B')
        self.dataset.group_name = 'Layer A'
        dataset_b.group_name = 'Layer B'
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(temp_dir, csv=False, png=True, plots=True)
            figure = ExportManager()._build_collection_plot_figure(
                'k_distribution', datasets, config,
            )

        ax = figure.axes[0]
        # The export reuses the shared spec's lognormal-histogram defaults
        # (frequency axis), matching the Comparison tab and the report.
        self.assertEqual(ax.get_title(), 'K Distribution (lognormal)')
        self.assertEqual(ax.get_ylabel(), 'Frequency (count)')
        self.assertGreater(len(ax.patches), 0)

    def test_k_distribution_is_a_selectable_collection_export(self):
        dataset_b = build_dataset('Sample B')
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir, csv=False, png=True, plots=True,
                selected_plot_types=['k_distribution'],
            )

            exported = ExportManager().export(datasets, config)

        basenames = {os.path.basename(path) for path in exported}
        self.assertEqual(len(exported), 1)
        self.assertTrue(any(name.endswith('_k_distribution.png') for name in basenames))

    def test_grouped_statistical_boxplot_uses_grouped_scope_series(self):
        dataset_b = build_dataset('Sample B')
        self.dataset.group_name = 'Layer A'
        dataset_b.group_name = 'Layer B'
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(temp_dir, csv=False, png=True, plots=True)
            with patch('gui.export_manager.render_k_scope_boxplot') as renderer:
                ExportManager()._build_collection_plot_figure(
                    'statistical_boxplots',
                    datasets,
                    config,
                )

        series = renderer.call_args[0][1]
        self.assertEqual([label for label, _values in series], ['Overall', 'Layer A', 'Layer B'])
        self.assertIn('Group', renderer.call_args.kwargs['title'])

    def test_progress_counts_actual_exported_files_not_export_batches(self):
        dataset_b = build_dataset('Sample B')
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv_long=True,
                csv_wide=True,
                excel=True,
                png=True,
                svg=True,
                plots=True,
                selected_plot_types=[
                    'grain_size_curve',
                    'k_value_bar',
                    'distribution_overlay',
                ],
            )
            progress = RecordingProgress()

            exported = ExportManager().export(datasets, config, progress)

            self.assertEqual(len(exported), 14)
            self.assertEqual(progress.maximum, len(exported))
            self.assertEqual(progress.values[-1], len(exported))

    def test_plot_export_uses_shared_renderer_and_plot_options(self):
        calls = []

        def capture_renderer(ax, particle_sizes, percent_passing, **kwargs):
            calls.append(kwargs)
            ax.plot(particle_sizes, percent_passing, label=kwargs.get('sample_name'))

        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv=False,
                png=True,
                plots=True,
                plot_include_legend=False,
                plot_include_grid=False,
                plot_contexts=[{
                    'style': PROFESSIONAL_STYLE,
                    'show_d_lines': True,
                    'show_markers': True,
                    'show_classification_zones': True,
                }],
            )

            with patch(
                'gui.export_manager.render_grain_size_distribution',
                side_effect=capture_renderer,
            ):
                exported = ExportManager().export(self.datasets, config)

            self.assertEqual(len(exported), 1)
            self.assertEqual(len(calls), 1)
            self.assertFalse(calls[0]['show_legend'])
            self.assertFalse(calls[0]['show_grid'])
            self.assertTrue(calls[0]['show_d_lines'])
            self.assertTrue(calls[0]['show_markers'])
            self.assertTrue(calls[0]['show_classification_zones'])

    def test_plot_export_keeps_live_legend_and_grid_state_when_allowed(self):
        calls = []

        def capture_renderer(ax, particle_sizes, percent_passing, **kwargs):
            calls.append(kwargs)
            ax.plot(particle_sizes, percent_passing, label=kwargs.get('sample_name'))

        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.make_config(
                temp_dir,
                csv=False,
                png=True,
                plots=True,
                plot_contexts=[{
                    'show_legend': False,
                    'show_grid': False,
                }],
            )

            with patch(
                'gui.export_manager.render_grain_size_distribution',
                side_effect=capture_renderer,
            ):
                exported = ExportManager().export(self.datasets, config)

        self.assertEqual(len(exported), 1)
        self.assertEqual(len(calls), 1)
        self.assertFalse(calls[0]['show_legend'])
        self.assertFalse(calls[0]['show_grid'])

    def test_plot_export_applies_axis_limits_from_plot_context(self):
        manager = ExportManager()

        figure = manager._build_grain_size_plot_figure(
            self.dataset.sample_name,
            self.dataset,
            self.make_config(output_dir='unused', csv=False),
            {
                'axis_limits': {
                    'xlim': (0.01, 10.0),
                    'ylim': (5.0, 95.0),
                },
            },
        )

        ax = figure.axes[0]
        self.assertAlmostEqual(ax.get_xlim()[0], 0.01)
        self.assertAlmostEqual(ax.get_xlim()[1], 10.0)
        self.assertAlmostEqual(ax.get_ylim()[0], 5.0)
        self.assertAlmostEqual(ax.get_ylim()[1], 95.0)
        figure.clear()

    def test_k_value_bar_figure_uses_canonical_m_s_unit(self):
        manager = ExportManager()
        results = build_results()

        figure = manager._build_k_value_bar_figure(
            'Sample A',
            results,
            self.make_config(output_dir='unused', csv=False),
            {'display_unit': HydraulicConductivityUnit.M_PER_DAY},
        )
        ax = figure.axes[0]

        self.assertIn('m/s', ax.get_ylabel())
        bar_heights = [patch.get_height() for patch in ax.patches]
        self.assertAlmostEqual(
            max(bar_heights), max(r.k_value for r in results), places=12
        )
        figure.clear()

    def test_single_sample_export_uses_global_style_over_context(self):
        # The global report style must win over any captured live-tab style, so
        # the Customize panel themes every export plot uniformly.
        import dataclasses
        from unittest.mock import patch
        from gui.plot_styles import PROFESSIONAL_STYLE

        sentinel = dataclasses.replace(PROFESSIONAL_STYLE, figure_facecolor='#123456')
        other = dataclasses.replace(PROFESSIONAL_STYLE, figure_facecolor='#abcdef')

        manager = ExportManager()
        with patch('gui.export_manager.resolve_report_style', return_value=sentinel):
            figure = manager._build_grain_size_plot_figure(
                self.dataset.sample_name, self.dataset,
                self.make_config(output_dir='unused', csv=False),
                {'style': other},  # captured live-tab style must be overridden
            )
        from matplotlib.colors import to_rgba
        self.assertEqual(figure.patch.get_facecolor(), to_rgba('#123456'))
        figure.clear()

    def test_collection_export_spec_uses_canonical_m_s_unit(self):
        from gui.plot_styles import PROFESSIONAL_STYLE

        dataset_b = build_dataset('Sample B')
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        spec = ExportManager()._build_export_comparison_spec(
            datasets,
            self.make_config(output_dir='unused', csv=False),
            {'display_unit': HydraulicConductivityUnit.M_PER_DAY},
            plot_type='k-values',
            style=PROFESSIONAL_STYLE,
        )
        self.assertEqual(spec.display_unit, HydraulicConductivityUnit.M_PER_S)

    def test_collection_export_spec_uses_global_palette(self):
        import gui.report_plot_style as rps
        from gui.plot_styles import PROFESSIONAL_STYLE

        dataset_b = build_dataset('Sample B')
        dataset_c = build_dataset('Sample C')
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
            (dataset_c.sample_name, dataset_c, self.results),
        ]
        config = self.make_config(output_dir='unused', csv=False)
        manager = ExportManager()
        try:
            rps.set_report_palette('Viridis')
            spec = manager._build_export_comparison_spec(
                datasets, config, None,
                plot_type='distribution', style=PROFESSIONAL_STYLE,
            )
            self.assertEqual(spec.palette, rps.resolve_report_palette_colors(3))
            # Viridis is not the categorical default, so the resolved colours differ.
            from gui.plot_constants import DATASET_COLORS
            self.assertNotEqual(spec.effective_colors[:3], DATASET_COLORS[:3])
        finally:
            rps.set_report_palette('Categorical')
            rps._reset_cache_for_tests()

    def test_export_k_bar_colors_follow_palette(self):
        import gui.report_plot_style as rps
        from gui.plot_constants import palette_colors
        from matplotlib.colors import to_hex

        manager = ExportManager()
        results = build_results()
        methods = [r.method_name for r in results]
        try:
            rps.set_report_palette('Viridis')
            figure = manager._build_k_value_bar_figure(
                'Sample A', results,
                self.make_config(output_dir='unused', csv=False), None,
            )
            colors = [to_hex(p.get_facecolor()) for p in figure.axes[0].patches]
            self.assertEqual(colors, palette_colors('Viridis', len(methods)))
        finally:
            rps.set_report_palette('Categorical')
            rps._reset_cache_for_tests()
            figure.clear()

    def test_collection_k_value_comparison_figure_uses_palette_colors(self):
        import gui.report_plot_style as rps
        from gui.plot_constants import DATASET_COLORS, palette_colors
        from matplotlib.colors import to_hex

        manager = ExportManager()
        dataset_b = build_dataset('Sample B')
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        config = self.make_config(output_dir='unused', csv=False)
        figure = None
        try:
            rps.set_report_palette('Viridis')
            figure = manager._build_collection_plot_figure(
                'k_value_comparison', datasets, config, None,
            )
            bar_colors = {to_hex(p.get_facecolor()) for p in figure.axes[0].patches}
            self.assertEqual(bar_colors, set(palette_colors('Viridis', 2)))
            self.assertNotEqual(bar_colors, set(DATASET_COLORS[:2]))
        finally:
            rps.set_report_palette('Categorical')
            rps._reset_cache_for_tests()
            if figure is not None:
                figure.clear()

    def test_ungrouped_statistical_boxplots_use_palette_colors(self):
        import gui.report_plot_style as rps
        from gui.plot_constants import palette_colors
        from matplotlib.colors import to_hex

        manager = ExportManager()
        dataset_b = build_dataset('Sample B')
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        config = self.make_config(output_dir='unused', csv=False)
        figure = None
        try:
            rps.set_report_palette('Viridis')
            snapshot = manager._build_comparison_snapshot_for_export(datasets, config)
            from k_aggregation import k_scope_value_series
            series = k_scope_value_series(snapshot.k)
            self.assertEqual(
                manager._k_scope_plot_colors(snapshot, series),
                palette_colors('Viridis', len(series)),
            )
            figure = manager._build_collection_plot_figure(
                'statistical_boxplots', datasets, config, None,
            )
            box_colors = [to_hex(p.get_facecolor()) for p in figure.axes[0].patches]
            self.assertEqual(box_colors, palette_colors('Viridis', len(series)))
        finally:
            rps.set_report_palette('Categorical')
            rps._reset_cache_for_tests()
            if figure is not None:
                figure.clear()

    def test_k_distribution_overall_scope_uses_spec_palette(self):
        import gui.report_plot_style as rps
        from gui.comparison_plot_spec import k_distribution_scopes
        from gui.plot_constants import palette_colors
        from gui.plot_styles import PROFESSIONAL_STYLE

        manager = ExportManager()
        dataset_b = build_dataset('Sample B')
        datasets = [
            (self.dataset.sample_name, self.dataset, self.results),
            (dataset_b.sample_name, dataset_b, self.results),
        ]
        config = self.make_config(output_dir='unused', csv=False)
        try:
            rps.set_report_palette('Viridis')
            spec = manager._build_export_comparison_spec(
                datasets, config, None,
                plot_type='k-distribution', style=PROFESSIONAL_STYLE,
            )
            scopes = k_distribution_scopes(spec)
            self.assertEqual(scopes[0]['label'], 'Overall')
            self.assertEqual(scopes[0]['color'], palette_colors('Viridis', 2)[0])
        finally:
            rps.set_report_palette('Categorical')
            rps._reset_cache_for_tests()

    def test_plot_export_applies_shared_text_options_from_context(self):
        manager = ExportManager()

        figure = manager._build_grain_size_plot_figure(
            self.dataset.sample_name,
            self.dataset,
            self.make_config(output_dir='unused', csv=False),
            {
                'show_title': False,
                'plot_title': 'Custom Title',
                'show_x_label': True,
                'plot_x_label': 'Custom X',
                'show_y_label': False,
                'plot_y_label': 'Custom Y',
            },
        )

        ax = figure.axes[0]
        self.assertEqual(ax.get_title(), '')
        self.assertEqual(ax.get_xlabel(), 'Custom X')
        self.assertEqual(ax.get_ylabel(), '')
        figure.clear()


class TestExportTabConfig(unittest.TestCase):
    def setUp(self):
        self.dataset = build_dataset()
        self.results = build_results()
        self.tab = ExportTab()
        self.tab.update_datasets(
            [(self.dataset.sample_name, self.dataset, self.results)],
            plot_contexts=[{'style': PROFESSIONAL_STYLE, 'show_grid': False}],
        )

    def tearDown(self):
        self.tab.deleteLater()


    def test_export_workspace_uses_two_panes_and_grouped_content_tabs(self):
        self.assertEqual(self.tab._export_splitter.count(), 2)
        tab_sets = []
        for tabs in self.tab.findChildren(QTabWidget):
            tab_sets.append({tabs.tabText(index) for index in range(tabs.count())})
        self.assertTrue(any(
            {'Data tables', 'Individual plots', 'Comparison plots'}.issubset(labels)
            for labels in tab_sets
        ))


    def test_closed_export_accordions_stack_without_internal_whitespace(self):
        self.tab.resize(430, 760)
        self.tab.show()
        APP.processEvents()

        self.tab.export_inspector_tabs.setCurrentIndex(0)
        self.tab.format_section.set_open(False)
        self.tab.content_section.set_open(False)
        APP.processEvents()

        format_bottom = self.tab.format_section.geometry().bottom()
        content_top = self.tab.content_section.geometry().top()

        self.assertLessEqual(self.tab.format_section.height(), 40)
        self.assertLessEqual(self.tab.content_section.height(), 40)
        self.assertLessEqual(content_top - format_bottom, 2)

    def test_plot_actions_are_contextual_to_selected_plots(self):
        self.assertFalse(hasattr(self.tab, 'plot_options_btn'))
        self.assertEqual(self.tab.open_dataset_btn.text(), 'Open Source Sample')
        self.assertTrue(self.tab.export_btn.text().startswith('Export'))

    def test_excel_format_defaults_to_combined_workbook(self):
        self.tab.selected_formats['excel'] = True
        self.tab.selected_formats['png'] = False
        config = self.tab._build_export_config()

        self.assertTrue(config['excel'])
        self.assertEqual(config['excel_mode'], 'combined')
        self.assertEqual(self.tab._estimate_export_file_count(), 4)

    def test_comparison_plot_breakdown_is_written_to_config(self):
        self.tab._toggle_content_item('plots', 'distribution_overlay', True)
        combo = self.tab.plot_breakdown_combos['distribution_overlay']
        combo.setCurrentIndex(combo.findData('both'))

        config = self.tab._build_export_config()

        self.assertEqual(config['plot_breakdowns']['distribution_overlay'], 'both')
        self.assertIn('distribution_overlay', config['selected_plot_types'])

    def test_k_distribution_plot_checkbox_is_exposed(self):
        self.assertIn('plots_k_distribution', self.tab.content_checkboxes)
        self.tab._toggle_content_item('plots', 'k_distribution', True)
        config = self.tab._build_export_config()
        self.assertIn('k_distribution', config['selected_plot_types'])

    def test_plot_content_options_are_written_to_export_config(self):
        self.tab._toggle_content_item('plots', 'include_legend', False)
        self.tab._toggle_content_item('plots', 'include_grid', False)
        self.tab._toggle_content_item('plots', 'k_value_bar', True)

        config = self.tab._build_export_config()

        self.assertFalse(config['plot_include_legend'])
        self.assertFalse(config['plot_include_grid'])
        self.assertEqual(config['selected_plot_types'], ['grain_size_curve', 'k_value_bar'])
        self.assertEqual(config['plot_contexts'][0]['style'], PROFESSIONAL_STYLE)
        self.assertFalse(config['plot_contexts'][0]['show_grid'])

    def test_disabling_grain_size_plot_item_disables_plot_export(self):
        self.tab._toggle_content_item('plots', 'grain_size_curve', False)

        config = self.tab._build_export_config()

        self.assertFalse(config['plots'])
        self.assertFalse(self.tab._plot_exports_enabled())
        self.assertNotIn('plot_contexts', config)

    def test_master_content_categories_gate_export_config(self):
        self.tab._toggle_category('grain_size', False)
        self.tab._toggle_category('k_values', False)
        self.tab._toggle_category('statistics', False)

        config = self.tab._build_export_config()

        self.assertFalse(config['grain_distribution'])
        self.assertFalse(config['percentiles'])
        self.assertFalse(config['gradation'])
        self.assertFalse(config['classification'])
        self.assertFalse(config['k_values'])
        self.assertFalse(config['statistics'])
        self.assertFalse(config['formulas'])
        self.assertFalse(config['validation'])
        self.assertEqual(config['selected_percentiles'], [])
        self.assertEqual(config['selected_k_categories'], {})
        self.assertEqual(config['selected_statistics'], [])
        self.assertFalse(config['include_grain_size_stats'])

    def test_percentile_checkbox_updates_config_and_csv_previews(self):
        self.tab.selected_formats.update({
            'csv_long': True,
            'csv_wide': True,
            'excel': False,
            'png': False,
            'svg': False,
            'pdf': False,
        })

        self.tab.content_checkboxes['percentile_d5'].setChecked(True)
        config = self.tab._build_export_config()

        self.assertIn('d5', config['selected_percentiles'])
        previews = {
            self.tab.preview_tabs.tabText(index): self.tab.preview_tabs.widget(index)
            for index in range(self.tab.preview_tabs.count())
        }
        long_preview = previews['CSV Long']
        wide_preview = previews['CSV Wide']
        long_headers = [
            long_preview.horizontalHeaderItem(column).text()
            for column in range(long_preview.columnCount())
        ]
        wide_headers = [
            wide_preview.horizontalHeaderItem(column).text()
            for column in range(wide_preview.columnCount())
        ]

        self.assertIn('D5 (mm)', long_headers)
        self.assertIn('D5_mm', wide_headers)

    def test_gradation_checkbox_updates_config_and_csv_previews(self):
        self.tab.selected_formats.update({
            'csv_long': True,
            'csv_wide': True,
            'excel': False,
            'png': False,
            'svg': False,
            'pdf': False,
        })

        self.tab.content_checkboxes['grain_size_gradation'].setChecked(False)
        config = self.tab._build_export_config()

        self.assertFalse(config['gradation'])
        previews = {
            self.tab.preview_tabs.tabText(index): self.tab.preview_tabs.widget(index)
            for index in range(self.tab.preview_tabs.count())
        }
        long_headers = [
            previews['CSV Long'].horizontalHeaderItem(column).text()
            for column in range(previews['CSV Long'].columnCount())
        ]
        wide_headers = [
            previews['CSV Wide'].horizontalHeaderItem(column).text()
            for column in range(previews['CSV Wide'].columnCount())
        ]

        self.assertNotIn('Cu', long_headers)
        self.assertNotIn('Cc', long_headers)
        self.assertNotIn('Cu_Uniformity_Coefficient', wide_headers)
        self.assertNotIn('Cc_Curvature_Coefficient', wide_headers)

    def test_classification_checkbox_updates_config_and_csv_previews(self):
        self.tab.selected_formats.update({
            'csv_long': True,
            'csv_wide': True,
            'excel': False,
            'png': False,
            'svg': False,
            'pdf': False,
        })

        self.tab.content_checkboxes['grain_size_classification'].setChecked(False)
        config = self.tab._build_export_config()

        self.assertFalse(config['classification'])
        previews = {
            self.tab.preview_tabs.tabText(index): self.tab.preview_tabs.widget(index)
            for index in range(self.tab.preview_tabs.count())
        }
        long_headers = [
            previews['CSV Long'].horizontalHeaderItem(column).text()
            for column in range(previews['CSV Long'].columnCount())
        ]
        wide_headers = [
            previews['CSV Wide'].horizontalHeaderItem(column).text()
            for column in range(previews['CSV Wide'].columnCount())
        ]

        self.assertNotIn('Soil Classification', long_headers)
        self.assertNotIn('Soil_Classification', wide_headers)

    def test_csv_previews_reflect_disabled_content_categories(self):
        self.tab.selected_formats.update({
            'csv_long': True,
            'csv_wide': True,
            'excel': False,
            'png': False,
            'svg': False,
            'pdf': False,
        })
        self.tab._toggle_category('grain_size', False)
        self.tab._toggle_category('k_values', False)
        self.tab._toggle_category('statistics', False)
        self.tab.update_preview()

        previews = {
            self.tab.preview_tabs.tabText(index): self.tab.preview_tabs.widget(index)
            for index in range(self.tab.preview_tabs.count())
        }
        long_preview = previews['CSV Long']
        wide_preview = previews['CSV Wide']

        self.assertIsInstance(long_preview, QTableWidget)
        self.assertIsInstance(wide_preview, QTableWidget)
        long_headers = [
            long_preview.horizontalHeaderItem(column).text()
            for column in range(long_preview.columnCount())
        ]
        wide_headers = [
            wide_preview.horizontalHeaderItem(column).text()
            for column in range(wide_preview.columnCount())
        ]

        self.assertNotIn('Method', long_headers)
        self.assertNotIn('D10 (mm)', long_headers)
        self.assertFalse(any(header.startswith('K ') for header in long_headers))
        self.assertNotIn('K_Hazen_m/s', wide_headers)
        self.assertNotIn('K_Mean_m/s', wide_headers)
        self.assertNotIn('D10_mm', wide_headers)

    def test_plot_preview_renders_canvas_and_can_request_dataset_jump(self):
        requested = []
        self.tab.jump_to_dataset_requested.connect(requested.append)

        self.tab.update_preview()
        self.assertIsNotNone(self.tab._plot_preview_table)
        self.assertEqual(self.tab._plot_preview_table.rowCount(), 1)
        self.assertIsNotNone(self.tab._plot_preview_canvas)

        self.tab._open_selected_plot_dataset()

        self.assertEqual(requested, [self.dataset.sample_name])

    def test_export_now_runs_on_worker_behind_loading_dialog(self):
        # Export is threaded: export_now builds a worker + cancellable
        # LoadingDialog and exec()s the dialog (the worker runs off the UI thread).
        with tempfile.TemporaryDirectory() as temp_dir:
            self.tab.output_dir.setText(temp_dir)
            fake_dialog = Mock()
            fake_worker = Mock()

            with patch('gui.export_tab.LoadingDialog', return_value=fake_dialog) as dialog_cls, \
                    patch('gui.export_worker.ExportWorker', return_value=fake_worker) as worker_cls:
                self.tab.export_now()

            dialog_cls.assert_called_once()
            # Dialog is cancellable now (cancel aborts the worker between files).
            self.assertTrue(dialog_cls.call_args.kwargs.get('cancellable'))
            worker_cls.assert_called_once()
            fake_worker.start.assert_called_once()
            fake_dialog.exec.assert_called_once()
            self.assertIsNotNone(self.tab._export_worker)

    def test_export_finished_shows_summary_and_clears_worker(self):
        # The success handler reports the file count and resets the re-entrancy
        # guard so a second export can run.
        self.tab._export_dialog = Mock()
        self.tab._export_worker = Mock()
        self.tab._export_output_dir = "out"
        with patch('gui.export_tab.QMessageBox.information') as info:
            self.tab._on_export_finished(["out/a.csv", "out/b.csv"])
        info.assert_called_once()
        self.assertIsNone(self.tab._export_worker)

    def test_plot_file_tree_groups_plot_files_by_dataset(self):
        other_dataset = build_dataset('Sample B')
        self.tab.update_datasets(
            [
                (self.dataset.sample_name, self.dataset, self.results),
                (other_dataset.sample_name, other_dataset, build_results()),
            ],
            plot_contexts=[
                {'style': PROFESSIONAL_STYLE, 'show_grid': False},
                {'style': PROFESSIONAL_STYLE, 'show_grid': True},
            ],
        )

        plot_folder = None
        for index in range(self.tab.file_tree.topLevelItemCount()):
            item = self.tab.file_tree.topLevelItem(index)
            if item.text(0) == 'plots':
                plot_folder = item
                break

        self.assertIsNotNone(plot_folder)
        self.assertEqual(plot_folder.childCount(), 2)
        self.assertEqual(plot_folder.child(0).text(0), self.dataset.sample_name)
        self.assertEqual(plot_folder.child(1).text(0), 'Sample B')
        self.assertGreater(plot_folder.child(0).childCount(), 0)
        self.assertEqual(plot_folder.child(0).child(0).text(0), 'plot.png')

    def test_grouped_plot_queue_selection_still_targets_dataset(self):
        other_dataset = build_dataset('Sample B')
        requested = []
        self.tab.jump_to_dataset_requested.connect(requested.append)
        self.tab.update_datasets(
            [
                (self.dataset.sample_name, self.dataset, self.results),
                (other_dataset.sample_name, other_dataset, build_results()),
            ],
            plot_contexts=[
                {'style': PROFESSIONAL_STYLE, 'show_grid': False},
                {'style': PROFESSIONAL_STYLE, 'show_grid': True},
            ],
        )

        tree = self.tab.plot_queue_tree
        self.assertEqual(tree.topLevelItemCount(), 2)
        self.assertEqual(tree.topLevelItem(1).text(0), 'Sample B')
        self.assertGreater(tree.topLevelItem(1).childCount(), 0)

        self.tab._select_plot_queue_row(1)
        self.assertEqual(tree.currentItem().parent().text(0), 'Sample B')
        self.tab._open_selected_plot_dataset()

        self.assertEqual(requested, ['Sample B'])

    def test_selected_scope_filters_datasets_and_plot_contexts(self):
        other_dataset = build_dataset('Sample B')
        other_results = build_results()
        self.tab.update_datasets(
            [
                (self.dataset.sample_name, self.dataset, self.results),
                (other_dataset.sample_name, other_dataset, other_results),
            ],
            plot_contexts=[
                {'style': PROFESSIONAL_STYLE, 'show_grid': False},
                {'style': PROFESSIONAL_STYLE, 'show_grid': True},
            ],
        )

        self.tab.selected_dataset_keys = {
            self.tab._dataset_key(other_dataset.sample_name, other_dataset)
        }
        self.tab.scope_selected.setChecked(True)

        datasets = self.tab._get_datasets_to_export()
        config = self.tab._build_export_config()

        self.assertEqual([name for name, _dataset, _results in datasets], ['Sample B'])
        self.assertEqual(len(config['plot_contexts']), 1)
        self.assertTrue(config['plot_contexts'][0]['show_grid'])

    def test_update_datasets_uses_selected_tabs_for_selected_scope(self):
        class StubDatasetTab:
            def __init__(self, dataset):
                self.dataset = dataset

            def get_dataset(self):
                return self.dataset

            def get_dataset_name(self):
                return self.dataset.sample_name

        other_dataset = build_dataset('Sample B')
        first_tab = StubDatasetTab(self.dataset)
        second_tab = StubDatasetTab(other_dataset)

        self.tab.update_datasets(
            [
                (self.dataset.sample_name, self.dataset, self.results),
                (other_dataset.sample_name, other_dataset, build_results()),
            ],
            plot_contexts=[
                {'style': PROFESSIONAL_STYLE, 'show_grid': False},
                {'style': PROFESSIONAL_STYLE, 'show_grid': True},
            ],
            dataset_tabs=[first_tab, second_tab],
            selected_tabs=[second_tab],
        )
        self.tab.scope_selected.setChecked(True)

        datasets = self.tab._get_datasets_to_export()
        config = self.tab._build_export_config()

        self.assertEqual([name for name, _dataset, _results in datasets], ['Sample B'])
        self.assertEqual(len(config['plot_contexts']), 1)
        self.assertTrue(config['plot_contexts'][0]['show_grid'])

    def test_dataset_scope_segments_expose_active_visual_state(self):
        self.tab.scope_selected.setChecked(True)

        self.assertFalse(self.tab.scope_all.property("active"))
        self.assertFalse(self.tab.scope_current.property("active"))
        self.assertTrue(self.tab.scope_selected.property("active"))
        self.assertIn(
            'QPushButton[exportScopeSeg="true"][active="true"]',
            self.tab.scope_segment_frame.styleSheet(),
        )

    def test_json_format_is_not_exposed_in_export_tab(self):
        format_keys = {
            button.property("format_key")
            for button in self.tab.findChildren(QPushButton)
            if button.property("format_key")
        }

        self.assertNotIn('json', self.tab.selected_formats)
        self.assertNotIn('json', format_keys)
        self.assertFalse(self.tab._build_export_config()['json'])


if __name__ == '__main__':
    unittest.main(verbosity=2)
