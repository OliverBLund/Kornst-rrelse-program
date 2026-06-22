"""
Export Manager - Handles all export operations for grain size analysis data
"""

import os
import csv
import json
from datetime import datetime
from typing import List, Dict, Optional, Any
from PyQt6.QtWidgets import QProgressDialog
from matplotlib.figure import Figure

from data_loader import GrainSizeData
from k_calculations_v2 import KCalculationResult
from analysis.comparison_snapshot import (
    ComparisonSnapshotOptions,
    DatasetAnalysisInput,
    build_comparison_snapshot,
)
from k_aggregation import KAggregationOptions, build_k_result_summary, k_scope_value_series
from grain_classification import ISO14688
from method_registry import DEFAULT_METHOD_ORDER, METHOD_CATEGORY_MAP, ordered_methods
from .plot_renderers import (
    apply_legend_aware_layout,
    render_applicability_heatmap,
    render_distribution_overlay,
    render_grain_size_distribution,
    render_k_bar_chart,
    render_k_overlay,
    render_k_scope_boxplot,
    render_reliability_matrix,
)
from .plot_context import (
    apply_axis_limits_from_context,
    grain_size_renderer_kwargs_from_context,
    plot_context_value,
    plot_style_from_context,
)
from .theme import apply_matplotlib_style


class ExportManager:
    """Manages export operations for various file formats"""

    DEFAULT_METHOD_ORDER = list(DEFAULT_METHOD_ORDER)
    METHOD_CATEGORY_MAP = METHOD_CATEGORY_MAP
    UNIT_SPECS = [
        ('m_s', 'K (m/s)', 'm/s', 'k_m_s', 1.0, '.3e'),
        ('cm_s', 'K (cm/s)', 'cm/s', 'k_cm_s', 100.0, '.3e'),
        ('m_d', 'K (m/d)', 'm/d', 'k_m_d', 86400.0, '.2f'),
    ]
    STAT_SPECS = [
        ('geometric_mean', 'Geometric Mean', 'K_Geometric_Mean', 'geometric_mean_k'),
        ('mean', 'Arithmetic Mean', 'K_Arithmetic_Mean', 'mean_k'),
        ('median', 'Median', 'K_Median', 'median_k'),
        ('std_dev', 'Std Dev', 'K_StdDev', 'stdev_k'),
        ('min', 'Min', 'K_Min', 'min_k'),
        ('max', 'Max', 'K_Max', 'max_k'),
        ('valid_count', 'Valid Count', 'Valid_Methods_Count', 'valid_count'),
    ]
    SINGLE_PLOT_TYPES = {
        'grain_size_curve',
        'k_value_bar',
        'applicability_heatmap',
    }
    COLLECTION_PLOT_TYPES = {
        'distribution_overlay',
        'k_value_comparison',
        'statistical_boxplots',
        'reliability_matrix',
    }
    PLOT_FILE_SUFFIXES = {
        'grain_size_curve': 'plot',
        'k_value_bar': 'k_values',
        'applicability_heatmap': 'applicability',
        'distribution_overlay': 'distribution_overlay',
        'k_value_comparison': 'k_value_comparison',
        'statistical_boxplots': 'k_value_boxplot',
        'reliability_matrix': 'reliability_matrix',
    }
    PERCENTILE_SPECS = [
        ('d5', 'D5', 5),
        ('d10', 'D10', 10),
        ('d16', 'D16', 16),
        ('d17', 'D17', 17),
        ('d20', 'D20', 20),
        ('d30', 'D30', 30),
        ('d50', 'D50', 50),
        ('d60', 'D60', 60),
        ('d84', 'D84', 84),
        ('d95', 'D95', 95),
    ]

    def __init__(self):
        self.exported_files = []
        self._scheme = ISO14688  # Active classification scheme; set via set_scheme()

    def set_scheme(self, scheme) -> None:
        """Set the active classification scheme used in all exports."""
        self._scheme = scheme

    def _get_enabled_unit_specs(self, config: Dict) -> List[tuple]:
        """Return enabled K-value units in a stable display order."""
        units = config.get('k_units') or {}
        return [spec for spec in self.UNIT_SPECS if units.get(spec[0], True)]

    def _get_primary_unit_spec(self, config: Dict) -> tuple:
        """Return the primary unit for compact summary sheets."""
        enabled = self._get_enabled_unit_specs(config)
        return enabled[0] if enabled else self.UNIT_SPECS[0]

    def _metadata_enabled(self, config: Dict, key: str) -> bool:
        """Return whether a metadata item should be exported."""
        metadata = config.get('include_metadata') or {}
        return metadata.get(key, True)

    def _get_export_timestamp(self, config: Dict) -> str:
        """Return the export timestamp shared across this export run."""
        return config.get('_export_timestamp', datetime.now().isoformat(timespec='seconds'))

    def _category_output_dir(self, config: Dict, *parts: str) -> str:
        """Return and create the output subdirectory for an export category."""
        base_dir = config['output_dir']
        if config.get('enforce_folder_structure', True) and parts:
            base_dir = os.path.join(base_dir, *parts)
        os.makedirs(base_dir, exist_ok=True)
        return base_dir

    def _selected_percentile_specs(self, config: Dict) -> List[tuple]:
        """Return selected percentile specs in UI/export order."""
        selected = config.get('selected_percentiles')
        if selected is None:
            selected = ['d10', 'd20', 'd30', 'd50', 'd60']
        selected_set = set(selected)
        return [spec for spec in self.PERCENTILE_SPECS if spec[0] in selected_set]

    def _percentile_value(self, dataset: GrainSizeData, percentile_num: int) -> Optional[float]:
        """Return a percentile value using dataset helpers when available."""
        getter = getattr(dataset, f'get_d{percentile_num}', None)
        if callable(getter):
            return getter()
        interpolator = getattr(dataset, '_interpolate_grain_size', None)
        if callable(interpolator):
            return interpolator(percentile_num)
        get_percentile = getattr(dataset, 'get_percentile', None)
        if callable(get_percentile):
            return get_percentile(percentile_num)
        return None

    def _effective_porosity(self, dataset: GrainSizeData) -> Optional[float]:
        """Return the porosity value used by calculations."""
        if hasattr(dataset, 'effective_porosity'):
            return dataset.effective_porosity()
        current = getattr(dataset, 'current_porosity', None)
        if current is not None:
            return current
        calculated = getattr(dataset, 'calculated_porosity', None)
        if calculated is not None:
            return calculated
        return getattr(dataset, 'porosity', None)

    def _porosity_source_label(self, dataset: GrainSizeData) -> str:
        if hasattr(dataset, 'porosity_source_label'):
            return dataset.porosity_source_label()
        return "Current dataset value"

    def _format_converted_value(self, k_value: float, unit_spec: tuple) -> str:
        """Format a conductivity value in the requested unit."""
        _, _, _, _, multiplier, fmt = unit_spec
        return format(k_value * multiplier, fmt)

    def _ordered_method_names(self, method_names) -> List[str]:
        """Return methods in the preferred domain order."""
        return ordered_methods(method_names, self.DEFAULT_METHOD_ORDER)

    def _selected_method_names(self, config: Dict) -> Optional[set[str]]:
        """Return selected method names or None when no filtering is requested."""
        filter_mode = config.get('k_filter_mode', 'all')

        if filter_mode == 'individual':
            return set(config.get('selected_k_methods') or [])

        if filter_mode == 'category':
            selected = set()
            for category, enabled in (config.get('selected_k_categories') or {}).items():
                if enabled:
                    selected.update(self.METHOD_CATEGORY_MAP.get(category, set()))
            return selected

        return None

    def _filter_results(self, results: Optional[List[KCalculationResult]], config: Dict) -> List[KCalculationResult]:
        """Filter K-results to the selected methods while preserving order."""
        filtered = [result for result in (results or []) if result.k_value is not None]
        selected = self._selected_method_names(config)
        if selected is None:
            return filtered
        return [result for result in filtered if result.method_name in selected]

    def _k_aggregation_options(self, config: Dict) -> KAggregationOptions:
        """Build shared K aggregation options from export settings."""
        return KAggregationOptions.from_methods(
            self._selected_method_names(config),
            include_warnings=bool(config.get('include_warning_k_statistics', False)),
            include_errors=bool(config.get('include_error_k_statistics', False)),
            method_order=tuple(self.DEFAULT_METHOD_ORDER),
        )

    def _build_comparison_snapshot_for_export(self, datasets: List[tuple], config: Dict):
        inputs = [
            DatasetAnalysisInput(
                label=str(name),
                dataset=dataset,
                k_results=tuple(results or ()),
                group_name=getattr(dataset, 'group_name', 'Ungrouped'),
            )
            for name, dataset, results in datasets
        ]
        return build_comparison_snapshot(
            inputs,
            ComparisonSnapshotOptions(
                k_options=self._k_aggregation_options(config),
                classification_scheme=self._scheme,
            ),
        )

    def _k_scope_plot_colors(self, snapshot, series) -> List[str]:
        uses_group_scope = any(group != 'Ungrouped' for group in snapshot.k.group_names)
        if not uses_group_scope:
            return []
        try:
            from .group_styles import group_color_map
            group_colors = group_color_map(snapshot.k.group_names)
        except Exception:
            fallback = ("#3a7ea0", "#6b8e23", "#b46428", "#2a9d8f", "#8b4580", "#a03a30")
            group_colors = {
                group: fallback[index % len(fallback)]
                for index, group in enumerate(snapshot.k.group_names)
            }
        return [
            "#8c6f45" if label == "Overall" else group_colors.get(label, "#777777")
            for label, _values in series
        ]

    def _collect_method_names(self, datasets: List[tuple], config: Dict) -> List[str]:
        """Collect method names visible in the current export selection."""
        method_names = []
        for _, _, results in datasets:
            method_names.extend(result.method_name for result in self._filter_results(results, config))

        if method_names:
            return self._ordered_method_names(method_names)

        selected = self._selected_method_names(config)
        if selected is not None:
            return self._ordered_method_names(selected)

        return list(self.DEFAULT_METHOD_ORDER)

    def _get_selected_stat_specs(self, config: Dict) -> List[tuple]:
        """Return selected statistics in a stable display order."""
        selected = config.get('selected_statistics')
        if selected is None:
            selected_names = {name for name, _, _, _ in self.STAT_SPECS}
        else:
            selected_names = set(selected)
        return [spec for spec in self.STAT_SPECS if spec[0] in selected_names]

    def _calculate_statistics(self, results: Optional[List[KCalculationResult]], config: Dict) -> Dict[str, float]:
        """Calculate the selected statistics from the filtered K-results."""
        filtered_results = self._filter_results(results, config)
        summary = build_k_result_summary(
            filtered_results,
            self._k_aggregation_options(config),
        )

        if summary.geometric_mean_m_s is None:
            return {}

        return {
            'geometric_mean_k': summary.geometric_mean_m_s,
            'mean_k': summary.arithmetic_mean_m_s,
            'median_k': summary.median_m_s,
            'stdev_k': summary.std_dev_m_s,
            'min_k': summary.min_m_s,
            'max_k': summary.max_m_s,
            'valid_count': summary.included_count,
        }

    def _append_dataset_context_columns(
        self,
        row: List[Any],
        dataset: GrainSizeData,
        config: Dict,
        percentile_specs: List[tuple],
        include_environmental: bool,
        include_timestamp: bool,
    ) -> None:
        """Append non-method dataset columns shared by CSV Long and Wide."""
        if include_environmental:
            row.extend([dataset.temperature, self._effective_porosity(dataset)])

        for _key, _label, percentile_num in percentile_specs:
            value = self._percentile_value(dataset, percentile_num)
            row.append(value if value is not None else '')

        if config.get('gradation', True):
            row.extend([
                dataset.get_uniformity_coefficient() if hasattr(dataset, 'get_uniformity_coefficient') else '',
                dataset.get_coefficient_of_curvature() if hasattr(dataset, 'get_coefficient_of_curvature') else '',
            ])

        if config.get('classification', True):
            row.append(dataset.classify(scheme=self._scheme).label)

        if include_timestamp:
            row.append(self._get_export_timestamp(config))

    def build_csv_long_table(
        self,
        datasets: List[tuple],
        config: Dict,
        max_data_rows: Optional[int] = None,
    ) -> List[List[Any]]:
        """Build CSV Long rows using the same schema as the exported file."""
        k_values_enabled = config.get('k_values', True)
        unit_specs = self._get_enabled_unit_specs(config) if k_values_enabled else []
        include_environmental = self._metadata_enabled(config, 'environmental')
        include_timestamp = self._metadata_enabled(config, 'export_timestamp')
        percentile_specs = self._selected_percentile_specs(config) if config.get('percentiles', True) else []

        header = ['Sample Name']
        if k_values_enabled:
            header.append('Method')
            header.extend(unit_label for _, unit_label, _, _, _, _ in unit_specs)
            header.append('Status')
            if config.get('validation', False):
                header.append('Status Message')
            if config.get('formulas', False):
                header.append('Formula')
        if include_environmental:
            header.extend(['Temperature (C)', 'Porosity'])
        header.extend(f'{label} (mm)' for _key, label, _num in percentile_specs)
        if config.get('gradation', True):
            header.extend(['Cu', 'Cc'])
        if config.get('classification', True):
            header.append('Soil Classification')
        if include_timestamp:
            header.append('Export Timestamp')

        rows: List[List[Any]] = [header]
        data_rows = 0

        for name, dataset, results in datasets:
            if k_values_enabled:
                for result in self._filter_results(results, config):
                    row = [name, result.method_name]
                    row.extend(self._format_converted_value(result.k_value, unit_spec) for unit_spec in unit_specs)
                    row.append(result.status.value if hasattr(result.status, 'value') else str(result.status))
                    if config.get('validation', False):
                        row.append(result.status_message)
                    if config.get('formulas', False):
                        row.append(result.formula_used)
                    self._append_dataset_context_columns(
                        row, dataset, config, percentile_specs, include_environmental, include_timestamp,
                    )
                    rows.append(row)
                    data_rows += 1
                    if max_data_rows is not None and data_rows >= max_data_rows:
                        return rows
            else:
                row = [name]
                self._append_dataset_context_columns(
                    row, dataset, config, percentile_specs, include_environmental, include_timestamp,
                )
                rows.append(row)
                data_rows += 1
                if max_data_rows is not None and data_rows >= max_data_rows:
                    return rows

        return rows

    def build_csv_wide_table(
        self,
        datasets: List[tuple],
        config: Dict,
        max_data_rows: Optional[int] = None,
    ) -> List[List[Any]]:
        """Build CSV Wide rows using the same schema as the exported file."""
        k_values_enabled = config.get('k_values', True)
        method_names = self._collect_method_names(datasets, config) if k_values_enabled else []
        unit_specs = self._get_enabled_unit_specs(config)
        include_environmental = self._metadata_enabled(config, 'environmental')
        include_timestamp = self._metadata_enabled(config, 'export_timestamp')
        selected_stats = self._get_selected_stat_specs(config) if config.get('statistics', True) else []
        percentile_specs = self._selected_percentile_specs(config) if config.get('percentiles', True) else []

        header = ['Sample_Name']
        if include_environmental:
            header.extend(['Temperature_C', 'Porosity'])

        for _key, label, _num in percentile_specs:
            header.append(f'{label}_mm')

        if config.get('gradation', True):
            header.extend(['Cu_Uniformity_Coefficient', 'Cc_Curvature_Coefficient'])

        if config.get('classification', True):
            header.append('Soil_Classification')

        if k_values_enabled:
            for method in method_names:
                safe_name = method.replace('-', '_').replace(' ', '_')
                for _, _, suffix, _, _, _ in unit_specs:
                    header.append(f'K_{safe_name}_{suffix}')

            for method in method_names:
                safe_name = method.replace('-', '_').replace(' ', '_')
                header.append(f'Status_{safe_name}')

        for _, _, suffix, _, _, _ in unit_specs:
            for _, _, header_prefix, value_key in selected_stats:
                if value_key != 'valid_count':
                    header.append(f'{header_prefix}_{suffix}')

        if any(value_key == 'valid_count' for _, _, _, value_key in selected_stats):
            header.append('Valid_Methods_Count')

        if include_timestamp:
            header.append('Export_Timestamp')

        rows: List[List[Any]] = [header]
        for row_index, (name, dataset, results) in enumerate(datasets):
            if max_data_rows is not None and row_index >= max_data_rows:
                break

            row = [name]
            if include_environmental:
                row.extend([dataset.temperature, self._effective_porosity(dataset)])

            for _key, _label, percentile_num in percentile_specs:
                value = self._percentile_value(dataset, percentile_num)
                row.append(f"{value:.4f}" if value is not None else '')

            if config.get('gradation', True):
                cu = dataset.get_uniformity_coefficient() if hasattr(dataset, 'get_uniformity_coefficient') else None
                cc = dataset.get_coefficient_of_curvature() if hasattr(dataset, 'get_coefficient_of_curvature') else None
                row.append(f"{cu:.2f}" if cu is not None else '')
                row.append(f"{cc:.2f}" if cc is not None else '')

            if config.get('classification', True):
                row.append(dataset.classify(scheme=self._scheme).label)

            if k_values_enabled:
                filtered_results = self._filter_results(results, config)
                method_results = {result.method_name: result for result in filtered_results}

                for method in method_names:
                    result = method_results.get(method)
                    for unit_spec in unit_specs:
                        row.append(self._format_converted_value(result.k_value, unit_spec) if result else '')

                for method in method_names:
                    result = method_results.get(method)
                    if result:
                        row.append(result.status.value if hasattr(result.status, 'value') else str(result.status))
                    else:
                        row.append('')

            stats_values = self._calculate_statistics(results, config)
            for unit_spec in unit_specs:
                for _, _, _, value_key in selected_stats:
                    if value_key != 'valid_count':
                        stat_value = stats_values.get(value_key)
                        row.append(self._format_converted_value(stat_value, unit_spec) if stat_value is not None else '')

            if any(value_key == 'valid_count' for _, _, _, value_key in selected_stats):
                count_value = stats_values.get('valid_count')
                row.append(str(count_value) if count_value is not None else '')

            if include_timestamp:
                row.append(self._get_export_timestamp(config))

            rows.append(row)

        return rows

    def export(self, datasets: List[tuple], config: Dict, progress: Optional[QProgressDialog] = None) -> List[str]:
        """
        Export datasets according to configuration

        Args:
            datasets: List of (name, GrainSizeData, List[KCalculationResult]) tuples
            config: Export configuration dictionary
            progress: Optional progress dialog

        Returns:
            List of exported file paths
        """
        self.exported_files = []
        config = dict(config)
        if self._metadata_enabled(config, 'export_timestamp') and '_export_timestamp' not in config:
            config['_export_timestamp'] = datetime.now().isoformat(timespec='seconds')
        total_steps = max(int(config.get('expected_file_count') or 0), self._calculate_total_steps(datasets, config))

        if progress:
            progress.setMaximum(total_steps)

        # CSV Export
        if config.get('csv', False):
            if config.get('csv_mode') == 'separate':
                if config.get('csv_long', True):
                    for name, dataset, results in datasets:
                        self._export_csv_single(name, dataset, results, config)
                        if progress:
                            progress.setValue(len(self.exported_files))
                if config.get('csv_wide', False):
                    self._export_csv_wide_format_filtered(datasets, config)
                    if progress:
                        progress.setValue(len(self.exported_files))
            else:
                if config.get('csv_long', True):
                    self._export_csv_combined_filtered(datasets, config)
                    if progress:
                        progress.setValue(len(self.exported_files))
                if config.get('csv_wide', False):
                    self._export_csv_wide_format_filtered(datasets, config)
                    if progress:
                        progress.setValue(len(self.exported_files))

        # Excel Export
        if config.get('excel', False):
            excel_mode = config.get('excel_mode', 'per_dataset')

            if excel_mode == 'per_dataset':
                for name, dataset, results in datasets:
                    self._export_excel_single(name, dataset, results, config)
                    if progress:
                        progress.setValue(len(self.exported_files))

            elif excel_mode == 'combined':
                self._export_excel_combined(datasets, config)
                if progress:
                    progress.setValue(len(self.exported_files))

            elif excel_mode == 'method_organized':
                self._export_excel_method_organized(datasets, config)
                if progress:
                    progress.setValue(len(self.exported_files))

        # JSON Export
        if config.get('json', False):
            for name, dataset, results in datasets:
                self._export_json(name, dataset, results, config)
                if progress:
                    progress.setValue(len(self.exported_files))

        # Plot Export
        selected_plot_types = self._selected_plot_types(config)
        if config.get('plots', False) and self._plot_formats_requested(config) and selected_plot_types:
            plot_contexts = config.get('plot_contexts') or []
            single_plot_types = [key for key in selected_plot_types if key in self.SINGLE_PLOT_TYPES]
            collection_plot_types = [key for key in selected_plot_types if key in self.COLLECTION_PLOT_TYPES]

            if single_plot_types:
                for idx, (name, dataset, results) in enumerate(datasets):
                    context = plot_contexts[idx] if idx < len(plot_contexts) else {}
                    self._export_single_sample_plots(name, dataset, results, config, context, single_plot_types)
                    if progress:
                        progress.setValue(len(self.exported_files))

            if collection_plot_types:
                self._export_collection_plots(datasets, config, plot_contexts, collection_plot_types)
                if progress:
                    progress.setValue(len(self.exported_files))

        return self.exported_files

    def _plot_formats_requested(self, config: Dict) -> bool:
        return any((
            config.get('png', False),
            config.get('svg', False),
            config.get('pdf_plot', False),
        ))

    def _selected_plot_types(self, config: Dict) -> List[str]:
        """Return selected plot types, preserving the legacy grain-plot default."""
        requested = config.get('selected_plot_types')
        if requested is None:
            requested = ['grain_size_curve'] if config.get('plots', False) else []

        allowed = self.SINGLE_PLOT_TYPES | self.COLLECTION_PLOT_TYPES
        selected = []
        for plot_type in requested:
            if plot_type in allowed and plot_type not in selected:
                selected.append(plot_type)
        return selected

    def _build_grain_size_plot_figure(
        self,
        name: str,
        dataset: GrainSizeData,
        config: Dict,
        context: Optional[Dict] = None,
    ) -> Figure:
        """Create a fresh plot figure through the shared renderer."""
        apply_matplotlib_style()
        style = plot_style_from_context(context)
        figure = Figure(figsize=config.get('plot_figsize', (10, 6)))
        figure.patch.set_facecolor(style.figure_facecolor)
        ax = figure.add_subplot(1, 1, 1)

        render_grain_size_distribution(
            ax,
            dataset.particle_sizes,
            dataset.percent_passing,
            **grain_size_renderer_kwargs_from_context(
                name,
                dataset,
                context,
                default_classification_scheme=self._scheme,
                include_grid=config.get('plot_include_grid', True),
                include_legend=config.get('plot_include_legend', True),
            ),
        )
        apply_axis_limits_from_context(ax, context)

        apply_legend_aware_layout(figure, style)
        return figure

    def _build_k_value_bar_figure(
        self,
        name: str,
        results: List[KCalculationResult],
        config: Dict,
        context: Optional[Dict] = None,
    ) -> Figure:
        apply_matplotlib_style()
        style = plot_style_from_context(context)
        figure = Figure(figsize=config.get('plot_figsize', (10, 6)))
        figure.patch.set_facecolor(style.figure_facecolor)
        ax = figure.add_subplot(1, 1, 1)

        filtered_results = self._filter_results(results, config)
        methods = [result.method_name for result in filtered_results]
        k_values = [result.k_value for result in filtered_results]
        flagged_methods = {
            result.method_name
            for result in filtered_results
            if not getattr(result, 'conditions_met', True)
            or getattr(getattr(result, 'status', None), 'name', 'OK') != 'OK'
        }
        reference_values = [
            result.k_value
            for result in filtered_results
            if result.method_name not in flagged_methods
            and result.k_value is not None
            and result.k_value > 0
        ]

        render_k_bar_chart(
            ax,
            methods,
            k_values,
            flagged_methods=flagged_methods,
            reference_values=reference_values,
            style=style,
            show_grid=config.get('plot_include_grid', True),
            show_legend=config.get('plot_include_legend', True),
            log_y_scale=bool(plot_context_value(context, 'log_k_y_scale', False)),
            sample_name=name,
        )
        apply_legend_aware_layout(figure, style)
        return figure

    def _build_applicability_heatmap_figure(
        self,
        name: str,
        results: List[KCalculationResult],
        config: Dict,
        context: Optional[Dict] = None,
    ) -> Figure:
        apply_matplotlib_style()
        style = plot_style_from_context(context)
        filtered_results = self._filter_results(results, config)
        figure = Figure(figsize=config.get('plot_figsize', (10, max(4, len(filtered_results) * 0.4))))
        figure.patch.set_facecolor(style.figure_facecolor)
        ax = figure.add_subplot(1, 1, 1)

        render_applicability_heatmap(
            ax,
            filtered_results,
            style=style,
            title=f"Method Applicability: {name}",
        )
        figure.tight_layout()
        return figure

    def _build_collection_plot_figure(
        self,
        plot_type: str,
        datasets: List[tuple],
        config: Dict,
        context: Optional[Dict] = None,
    ) -> Figure:
        apply_matplotlib_style()
        style = plot_style_from_context(context)
        figure = Figure(figsize=config.get('plot_figsize', (12, 7)))
        figure.patch.set_facecolor(style.figure_facecolor)
        ax = figure.add_subplot(1, 1, 1)

        if plot_type == 'distribution_overlay':
            render_distribution_overlay(
                ax,
                [dataset for _, dataset, _ in datasets],
                labels=[name for name, _, _ in datasets],
                style=style,
                show_grid=config.get('plot_include_grid', True),
                show_legend=config.get('plot_include_legend', True),
            )
            apply_legend_aware_layout(figure, style)
            return figure

        if plot_type == 'k_value_comparison':
            k_results_dict = {}
            flagged_methods_dict = {}
            for name, _, results in datasets:
                filtered_results = self._filter_results(results, config)
                k_results_dict[name] = {
                    result.method_name: result.k_value
                    for result in filtered_results
                    if result.k_value is not None
                }
                flagged_methods_dict[name] = {
                    result.method_name
                    for result in filtered_results
                    if not getattr(result, 'conditions_met', True)
                    or getattr(getattr(result, 'status', None), 'name', 'OK') != 'OK'
                }
            render_k_overlay(
                ax,
                k_results_dict,
                flagged_methods_dict=flagged_methods_dict,
                style=style,
                show_grid=config.get('plot_include_grid', True),
                show_legend=config.get('plot_include_legend', True),
                log_y_scale=bool(plot_context_value(context, 'log_k_y_scale', False)),
            )
            apply_legend_aware_layout(figure, style)
            return figure

        if plot_type == 'statistical_boxplots':
            snapshot = self._build_comparison_snapshot_for_export(datasets, config)
            series = k_scope_value_series(snapshot.k)
            uses_group_scope = any(group != 'Ungrouped' for group in snapshot.k.group_names)
            render_k_scope_boxplot(
                ax,
                series,
                colors=self._k_scope_plot_colors(snapshot, series),
                style=style,
                show_grid=config.get('plot_include_grid', True),
                title=(
                    "Hydraulic Conductivity Distribution by Group"
                    if uses_group_scope
                    else "Hydraulic Conductivity Distribution by Dataset"
                ),
            )
            figure.tight_layout()
            return figure

        if plot_type == 'reliability_matrix':
            render_reliability_matrix(
                ax,
                {name: self._filter_results(results, config) for name, _, results in datasets},
                style=style,
            )
            figure.tight_layout()
            return figure

        raise ValueError(f"Unsupported plot type: {plot_type}")

    def _build_single_sample_plot_figure(
        self,
        plot_type: str,
        name: str,
        dataset: GrainSizeData,
        results: List[KCalculationResult],
        config: Dict,
        context: Optional[Dict] = None,
    ) -> Figure:
        if plot_type == 'grain_size_curve':
            return self._build_grain_size_plot_figure(name, dataset, config, context)
        if plot_type == 'k_value_bar':
            return self._build_k_value_bar_figure(name, results, config, context)
        if plot_type == 'applicability_heatmap':
            return self._build_applicability_heatmap_figure(name, results, config, context)
        raise ValueError(f"Unsupported plot type: {plot_type}")

    def _save_plot_figure(self, figure: Figure, base_filename: str, plot_type: str, config: Dict) -> None:
        output_dir = self._category_output_dir(
            config,
            'plots',
            config.get('_active_plot_folder', 'collection'),
        )
        suffix = self.PLOT_FILE_SUFFIXES.get(plot_type, plot_type)
        figure.patch.set_facecolor('white')
        figure.patch.set_alpha(1.0)

        if config.get('png', False):
            filepath = os.path.join(output_dir, f"{base_filename}_{suffix}.png")
            figure.savefig(
                filepath,
                dpi=config.get('png_dpi', 300),
                bbox_inches='tight',
                facecolor='white',
                edgecolor='white',
            )
            self.exported_files.append(filepath)

        if config.get('svg', False):
            filepath = os.path.join(output_dir, f"{base_filename}_{suffix}.svg")
            figure.savefig(filepath, format='svg', bbox_inches='tight', facecolor='white', edgecolor='white')
            self.exported_files.append(filepath)

        if config.get('pdf_plot', False):
            filepath = os.path.join(output_dir, f"{base_filename}_{suffix}.pdf")
            figure.savefig(filepath, format='pdf', bbox_inches='tight', facecolor='white', edgecolor='white')
            self.exported_files.append(filepath)

    def _export_single_sample_plots(
        self,
        name: str,
        dataset: GrainSizeData,
        results: List[KCalculationResult],
        config: Dict,
        context: Optional[Dict] = None,
        plot_types: Optional[List[str]] = None,
    ):
        """Export selected single-sample plot files."""
        template = config['filename_template']
        base_filename = self._format_filename(template, name, '')
        previous_folder = config.get('_active_plot_folder')
        config['_active_plot_folder'] = self._format_filename('{sample_name}', name, '')

        try:
            for plot_type in plot_types or ['grain_size_curve']:
                figure = self._build_single_sample_plot_figure(
                    plot_type, name, dataset, results, config, context,
                )
                self._save_plot_figure(figure, base_filename, plot_type, config)
                figure.clear()
        finally:
            if previous_folder is None:
                config.pop('_active_plot_folder', None)
            else:
                config['_active_plot_folder'] = previous_folder

    def _export_plots(
        self,
        name: str,
        dataset: GrainSizeData,
        results: List[KCalculationResult],
        config: Dict,
        context: Optional[Dict] = None,
    ):
        """Legacy entry point: export the grain-size curve plot."""
        self._export_single_sample_plots(
            name, dataset, results, config, context, ['grain_size_curve'],
        )

    def _export_collection_plots(
        self,
        datasets: List[tuple],
        config: Dict,
        plot_contexts: Optional[List[Dict]] = None,
        plot_types: Optional[List[str]] = None,
    ) -> None:
        """Export selected collection-level comparison plots."""
        template = config['filename_template']
        collection_name = config.get('collection_sample_name', 'all_datasets')
        base_filename = self._format_filename(template, collection_name, '')
        context = plot_contexts[0] if plot_contexts else {}
        previous_folder = config.get('_active_plot_folder')
        config['_active_plot_folder'] = self._format_filename('{sample_name}', collection_name, '')

        try:
            for plot_type in plot_types or []:
                figure = self._build_collection_plot_figure(plot_type, datasets, config, context)
                self._save_plot_figure(figure, base_filename, plot_type, config)
                figure.clear()
        finally:
            if previous_folder is None:
                config.pop('_active_plot_folder', None)
            else:
                config['_active_plot_folder'] = previous_folder

    def _calculate_total_steps(self, datasets: List[tuple], config: Dict) -> int:
        """Calculate expected output file count for progress reporting."""
        steps = 0

        if config.get('csv', False):
            if config.get('csv_mode') == 'separate':
                if config.get('csv_long', True):
                    for _name, _dataset, results in datasets:
                        if config.get('grain_distribution', True):
                            steps += 1
                        if config.get('k_values', True) and results:
                            steps += 1
                        if config.get('statistics', True):
                            steps += 1
                if config.get('csv_wide', False):
                    steps += 1
            else:
                if config.get('csv_long', True):
                    steps += 1
                if config.get('csv_wide', False):
                    steps += 1

        if config.get('excel', False):
            excel_mode = config.get('excel_mode', 'per_dataset')
            if excel_mode == 'per_dataset':
                steps += len(datasets)
            else:
                steps += 1

        if config.get('json', False):
            steps += len(datasets)

        if config.get('plots', False) and self._plot_formats_requested(config):
            selected_plot_types = self._selected_plot_types(config)
            plot_format_count = sum(
                1
                for key in ('png', 'svg', 'pdf_plot')
                if config.get(key, False)
            )
            single_count = len([plot_type for plot_type in selected_plot_types if plot_type in self.SINGLE_PLOT_TYPES])
            collection_count = len([plot_type for plot_type in selected_plot_types if plot_type in self.COLLECTION_PLOT_TYPES])
            steps += ((len(datasets) * single_count) + collection_count) * plot_format_count

        return max(steps, 1)

    def _format_filename(self, template: str, name: str, extension: str = "") -> str:
        """Format filename from template"""
        now = datetime.now()

        replacements = {
            '{sample_name}': name,
            '{date}': now.strftime('%Y%m%d'),
            '{time}': now.strftime('%H%M%S'),
            '{project}': 'grain_analysis',
            '{method}': 'all'
        }

        filename = template
        for key, value in replacements.items():
            filename = filename.replace(key, value)

        # Remove any invalid filename characters
        invalid_chars = '<>:"|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')

        if extension and not filename.endswith(extension):
            filename += extension

        return filename

    # ==================== CSV EXPORT ====================

    def _export_csv_single(self, name: str, dataset: GrainSizeData,
                          results: List[KCalculationResult], config: Dict):
        """Export single dataset to CSV"""
        output_dir = self._category_output_dir(config, 'tables', 'csv', self._format_filename('{sample_name}', name, ''))
        template = config['filename_template']

        # Generate base filename
        base_filename = self._format_filename(template, name, '')

        # Export grain size data if requested
        if config.get('grain_distribution', True):
            filename = f"{base_filename}_grain_size.csv"
            filepath = os.path.join(output_dir, filename)
            self._write_grain_size_csv(filepath, dataset)
            self.exported_files.append(filepath)

        # Export K-values if requested
        if config.get('k_values', True) and results:
            filename = f"{base_filename}_k_values.csv"
            filepath = os.path.join(output_dir, filename)
            self._write_k_values_csv_filtered(filepath, name, dataset, results, config)
            self.exported_files.append(filepath)

        # Export statistics if requested
        if config.get('statistics', True):
            filename = f"{base_filename}_statistics.csv"
            filepath = os.path.join(output_dir, filename)
            self._write_statistics_csv_filtered(filepath, dataset, results, config)
            self.exported_files.append(filepath)

    def _export_csv_combined(self, datasets: List[tuple], config: Dict):
        """Export all datasets to a single combined CSV"""
        output_dir = self._category_output_dir(config, 'workbooks')
        template = config['filename_template']

        filename = self._format_filename(template, 'combined_all_datasets', '.csv')
        filepath = os.path.join(output_dir, filename)

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Write header
            writer.writerow([
                'Sample Name', 'Method', 'K (m/s)', 'K (cm/s)', 'K (m/d)',
                'Status', 'Formula', 'Temperature (°C)', 'Porosity',
                'D10 (mm)', 'D50 (mm)', 'D60 (mm)', 'Cu', 'Cc'
            ])

            # Write data for each dataset
            for name, dataset, results in datasets:
                if results:
                    for result in results:
                        if result.k_value is not None:
                            k_cm_s = result.k_value * 100
                            k_m_d = result.k_value * 86400

                            writer.writerow([
                                name,
                                result.method_name,
                                f"{result.k_value:.3e}",
                                f"{k_cm_s:.3e}",
                                f"{k_m_d:.2f}",
                                result.status.value if hasattr(result.status, 'value') else str(result.status),
                                result.formula_used,
                                result.temperature,
                                result.porosity,
                                dataset.get_d10() if hasattr(dataset, 'get_d10') else '',
                                dataset.get_d50() if hasattr(dataset, 'get_d50') else '',
                                dataset.get_d60() if hasattr(dataset, 'get_d60') else '',
                                dataset.get_uniformity_coefficient() if hasattr(dataset, 'get_uniformity_coefficient') else '',
                                dataset.get_coefficient_of_curvature() if hasattr(dataset, 'get_coefficient_of_curvature') else ''
                            ])

        self.exported_files.append(filepath)


    def _export_csv_wide_format(self, datasets: List[tuple], config: Dict):
        """
        Export wide-format CSV with one row per dataset and columns for all parameters.
        Perfect for statistical analysis and comparison.
        """
        output_dir = self._category_output_dir(config, 'workbooks')
        template = config['filename_template']

        filename = self._format_filename(template, 'wide_format_all_datasets', '.csv')
        filepath = os.path.join(output_dir, filename)

        # Get all unique method names from results (dynamically)
        all_method_names = set()
        for name, dataset, results in datasets:
            if results:
                for result in results:
                    all_method_names.add(result.method_name)

        # Sort method names for consistent ordering
        method_names = self._ordered_method_names(all_method_names)

        # If no results yet, use standard methods as fallback
        if not method_names:
            method_names = list(self.DEFAULT_METHOD_ORDER)

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Build header DYNAMICALLY based on actual method names
            # Get selected percentiles from config (or use defaults)
            selected_percentiles = config.get('selected_percentiles', ['d10', 'd20', 'd30', 'd50', 'd60'])
            percentile_mapping = {
                'd5': 5, 'd10': 10, 'd16': 16, 'd17': 17, 'd20': 20,
                'd30': 30, 'd50': 50, 'd60': 60, 'd84': 84, 'd95': 95
            }

            header = [
                # Sample info
                'Sample_Name',
                'Temperature_C',
                'Porosity',
            ]

            # Add only selected grain size percentiles
            for p_key in selected_percentiles:
                p_num = percentile_mapping.get(p_key, 0)
                if p_num > 0:
                    header.append(f'D{p_num}_mm')

            # Gradation parameters
            if config.get('gradation', True):
                header.extend([
                    'Cu_Uniformity_Coefficient',
                    'Cc_Curvature_Coefficient',
                ])

            # Add K-values for each method in m/s
            for method in method_names:
                # Replace special chars for column names
                safe_name = method.replace('-', '_').replace(' ', '_')
                header.append(f'K_{safe_name}_m/s')

            # Add K-values in cm/s
            for method in method_names:
                safe_name = method.replace('-', '_').replace(' ', '_')
                header.append(f'K_{safe_name}_cm/s')

            # Add K-values in m/d
            for method in method_names:
                safe_name = method.replace('-', '_').replace(' ', '_')
                header.append(f'K_{safe_name}_m/d')

            # Add status flags for each method
            for method in method_names:
                safe_name = method.replace('-', '_').replace(' ', '_')
                header.append(f'Status_{safe_name}')

            # Add statistical summaries
            header.extend([
                'K_Mean_m/s', 'K_Median_m/s', 'K_StdDev_m/s', 'K_Min_m/s', 'K_Max_m/s',
                'K_Mean_cm/s', 'K_Median_cm/s', 'K_Min_cm/s', 'K_Max_cm/s',
                'K_Mean_m/d', 'K_Median_m/d', 'K_Min_m/d', 'K_Max_m/d',
                'Valid_Methods_Count'
            ])

            writer.writerow(header)

            # Write data for each dataset
            for name, dataset, results in datasets:
                row = []

                # Sample info
                row.append(name)
                row.append(dataset.temperature)
                row.append(self._effective_porosity(dataset))

                # Grain size percentiles - only selected ones
                for p_key in selected_percentiles:
                    p_num = percentile_mapping.get(p_key, 0)
                    if p_num > 0 and hasattr(dataset, '_interpolate_grain_size'):
                        value = dataset._interpolate_grain_size(p_num)
                        row.append(f"{value:.4f}" if value is not None else '')
                    elif hasattr(dataset, f'get_d{p_num}'):
                        value = getattr(dataset, f'get_d{p_num}')()
                        row.append(f"{value:.4f}" if value is not None else '')
                    else:
                        row.append('')

                # Gradation parameters (if enabled in config)
                if config.get('gradation', True):
                    cu = dataset.get_uniformity_coefficient() if hasattr(dataset, 'get_uniformity_coefficient') else None
                    cc = dataset.get_coefficient_of_curvature() if hasattr(dataset, 'get_coefficient_of_curvature') else None
                    row.append(f"{cu:.2f}" if cu is not None else '')
                    row.append(f"{cc:.2f}" if cc is not None else '')

                # Build method -> result mapping
                method_results = {}
                if results:
                    for result in results:
                        method_results[result.method_name] = result

                # K-values in m/s
                for method in method_names:
                    if method in method_results:
                        k_val = method_results[method].k_value
                        row.append(f"{k_val:.3e}" if k_val is not None else '')
                    else:
                        row.append('')

                # K-values in cm/s
                for method in method_names:
                    if method in method_results:
                        k_val = method_results[method].k_value
                        row.append(f"{k_val * 100:.3e}" if k_val is not None else '')
                    else:
                        row.append('')

                # K-values in m/d
                for method in method_names:
                    if method in method_results:
                        k_val = method_results[method].k_value
                        row.append(f"{k_val * 86400:.2f}" if k_val is not None else '')
                    else:
                        row.append('')

                # Status flags
                for method in method_names:
                    if method in method_results:
                        status = method_results[method].status
                        row.append(status.value if hasattr(status, 'value') else str(status))
                    else:
                        row.append('')

                # Statistical summaries
                valid_k_values = [r.k_value for r in results if r.k_value is not None and r.k_value > 0] if results else []

                if valid_k_values:
                    import statistics

                    mean_k = statistics.mean(valid_k_values)
                    median_k = statistics.median(valid_k_values)
                    stdev_k = statistics.stdev(valid_k_values) if len(valid_k_values) > 1 else 0
                    min_k = min(valid_k_values)
                    max_k = max(valid_k_values)

                    # m/s
                    row.append(f"{mean_k:.3e}")
                    row.append(f"{median_k:.3e}")
                    row.append(f"{stdev_k:.3e}")
                    row.append(f"{min_k:.3e}")
                    row.append(f"{max_k:.3e}")

                    # cm/s
                    row.append(f"{mean_k * 100:.3e}")
                    row.append(f"{median_k * 100:.3e}")
                    row.append(f"{min_k * 100:.3e}")
                    row.append(f"{max_k * 100:.3e}")

                    # m/d
                    row.append(f"{mean_k * 86400:.2f}")
                    row.append(f"{median_k * 86400:.2f}")
                    row.append(f"{min_k * 86400:.2f}")
                    row.append(f"{max_k * 86400:.2f}")

                    row.append(len(valid_k_values))
                else:
                    # Empty statistics
                    row.extend([''] * 13)

                writer.writerow(row)

        self.exported_files.append(filepath)

    def _write_grain_size_csv(self, filepath: str, dataset: GrainSizeData):
        """Write grain size distribution to CSV"""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Header
            writer.writerow(['Particle Size (mm)', 'Percent Passing (%)'])

            # Data
            for size, percent in zip(dataset.particle_sizes, dataset.percent_passing):
                writer.writerow([f"{size:.4f}", f"{percent:.2f}"])

    def _write_k_values_csv(self, filepath: str, name: str, dataset: GrainSizeData,
                           results: List[KCalculationResult], config: Dict):
        """Write K-values to CSV"""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Header
            header = ['Method', 'K (m/s)', 'K (cm/s)', 'K (m/d)', 'Status', 'Status Message']

            if config.get('formulas', False):
                header.append('Formula')

            header.extend(['Temperature (°C)', 'Porosity', 'Grain Size Used'])

            writer.writerow(header)

            # Data
            for result in results:
                if result.k_value is not None:
                    k_cm_s = result.k_value * 100
                    k_m_d = result.k_value * 86400

                    row = [
                        result.method_name,
                        f"{result.k_value:.3e}",
                        f"{k_cm_s:.3e}",
                        f"{k_m_d:.2f}",
                        result.status.value if hasattr(result.status, 'value') else str(result.status),
                        result.status_message
                    ]

                    if config.get('formulas', False):
                        row.append(result.formula_used)

                    row.extend([
                        result.temperature,
                        result.porosity,
                        result.grain_size_used
                    ])

                    writer.writerow(row)

    def _write_statistics_csv(self, filepath: str, dataset: GrainSizeData,
                             results: List[KCalculationResult], config: Dict):
        """Write statistics to CSV"""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Sample Information
            writer.writerow(['Sample Information'])
            writer.writerow(['Parameter', 'Value'])
            writer.writerow(['Sample Name', dataset.sample_name])
            writer.writerow(['Temperature (°C)', dataset.temperature])
            writer.writerow(['Porosity', self._effective_porosity(dataset)])
            writer.writerow(['Porosity Source', self._porosity_source_label(dataset)])
            writer.writerow([])

            # Grain Size Percentiles
            if config.get('percentiles', True):
                writer.writerow(['Grain Size Percentiles'])
                writer.writerow(['Percentile', 'Size (mm)'])

                percentiles = {
                    'D10': dataset.get_d10() if hasattr(dataset, 'get_d10') else None,
                    'D20': dataset.get_d20() if hasattr(dataset, 'get_d20') else None,
                    'D30': dataset.get_d30() if hasattr(dataset, 'get_d30') else None,
                    'D50': dataset.get_d50() if hasattr(dataset, 'get_d50') else None,
                    'D60': dataset.get_d60() if hasattr(dataset, 'get_d60') else None,
                }

                for name, value in percentiles.items():
                    if value is not None:
                        writer.writerow([name, f"{value:.4f}"])

                writer.writerow([])

            # Gradation Parameters
            if config.get('gradation', True):
                writer.writerow(['Gradation Parameters'])
                writer.writerow(['Parameter', 'Value'])

                cu = dataset.get_uniformity_coefficient() if hasattr(dataset, 'get_uniformity_coefficient') else None
                cc = dataset.get_coefficient_of_curvature() if hasattr(dataset, 'get_coefficient_of_curvature') else None

                if cu is not None:
                    writer.writerow(['Uniformity Coefficient (Cu)', f"{cu:.2f}"])
                if cc is not None:
                    writer.writerow(['Coefficient of Curvature (Cc)', f"{cc:.2f}"])

                writer.writerow([])

            # K-Value Statistics
            if config.get('statistics', True) and results:
                valid_k_values = [r.k_value for r in results if r.k_value is not None and r.k_value > 0]

                if valid_k_values:
                    import statistics

                    writer.writerow(['K-Value Statistics'])
                    writer.writerow(['Statistic', 'Value (m/s)', 'Value (cm/s)', 'Value (m/d)'])

                    mean_k = statistics.mean(valid_k_values)
                    median_k = statistics.median(valid_k_values)
                    min_k = min(valid_k_values)
                    max_k = max(valid_k_values)

                    if len(valid_k_values) > 1:
                        stdev_k = statistics.stdev(valid_k_values)
                    else:
                        stdev_k = 0

                    writer.writerow(['Mean', f"{mean_k:.3e}", f"{mean_k*100:.3e}", f"{mean_k*86400:.2f}"])
                    writer.writerow(['Median', f"{median_k:.3e}", f"{median_k*100:.3e}", f"{median_k*86400:.2f}"])
                    writer.writerow(['Std Dev', f"{stdev_k:.3e}", f"{stdev_k*100:.3e}", f"{stdev_k*86400:.2f}"])
                    writer.writerow(['Min', f"{min_k:.3e}", f"{min_k*100:.3e}", f"{min_k*86400:.2f}"])
                    writer.writerow(['Max', f"{max_k:.3e}", f"{max_k*100:.3e}", f"{max_k*86400:.2f}"])

    def _export_csv_combined_filtered(self, datasets: List[tuple], config: Dict):
        """Export a filtered long-format CSV honoring method and unit selections."""
        output_dir = self._category_output_dir(config, 'tables', 'csv')
        template = config['filename_template']

        filename = self._format_filename(template, 'combined_all_datasets', '.csv')
        filepath = os.path.join(output_dir, filename)

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(self.build_csv_long_table(datasets, config))

        self.exported_files.append(filepath)

    def _export_csv_wide_format_filtered(self, datasets: List[tuple], config: Dict):
        """Export a filtered wide-format CSV honoring method and unit selections."""
        output_dir = self._category_output_dir(config, 'tables', 'csv')
        template = config['filename_template']

        filename = self._format_filename(template, 'wide_format_all_datasets', '.csv')
        filepath = os.path.join(output_dir, filename)

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(self.build_csv_wide_table(datasets, config))

        self.exported_files.append(filepath)

    def _write_k_values_csv_filtered(self, filepath: str, name: str, dataset: GrainSizeData,
                                     results: List[KCalculationResult], config: Dict):
        """Write filtered K-values to CSV honoring method and unit selection."""
        unit_specs = self._get_enabled_unit_specs(config)
        filtered_results = self._filter_results(results, config)
        include_environmental = self._metadata_enabled(config, 'environmental')
        include_timestamp = self._metadata_enabled(config, 'export_timestamp')

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            header = ['Method']
            header.extend(unit_label for _, unit_label, _, _, _, _ in unit_specs)
            header.append('Status')
            if config.get('validation', False):
                header.append('Status Message')
            if config.get('formulas', False):
                header.append('Formula')
            if include_environmental:
                header.extend(['Temperature (C)', 'Porosity'])
            header.append('Grain Size Used')
            if include_timestamp:
                header.append('Export Timestamp')
            writer.writerow(header)

            for result in filtered_results:
                row = [result.method_name]
                row.extend(self._format_converted_value(result.k_value, unit_spec) for unit_spec in unit_specs)
                row.append(result.status.value if hasattr(result.status, 'value') else str(result.status))
                if config.get('validation', False):
                    row.append(result.status_message)
                if config.get('formulas', False):
                    row.append(result.formula_used)
                if include_environmental:
                    row.extend([result.temperature, result.porosity])
                row.append(result.grain_size_used)
                if include_timestamp:
                    row.append(self._get_export_timestamp(config))
                writer.writerow(row)

    def _write_statistics_csv_filtered(self, filepath: str, dataset: GrainSizeData,
                                       results: List[KCalculationResult], config: Dict):
        """Write statistics CSV using the filtered K-result set."""
        unit_specs = self._get_enabled_unit_specs(config)
        filtered_results = self._filter_results(results, config)
        include_sample_info = self._metadata_enabled(config, 'sample_info')
        include_environmental = self._metadata_enabled(config, 'environmental')
        include_timestamp = self._metadata_enabled(config, 'export_timestamp')
        include_grain_size_stats = config.get('include_grain_size_stats', True)
        selected_stats = self._get_selected_stat_specs(config)
        stats_values = self._calculate_statistics(results, config)

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            if include_sample_info or include_environmental or include_timestamp:
                writer.writerow(['Sample Information'])
                writer.writerow(['Parameter', 'Value'])
                if include_sample_info:
                    writer.writerow(['Sample Name', dataset.sample_name])
                if include_environmental:
                    writer.writerow(['Temperature (C)', dataset.temperature])
                    writer.writerow(['Porosity', self._effective_porosity(dataset)])
                    writer.writerow(['Porosity Source', self._porosity_source_label(dataset)])
                if include_timestamp:
                    writer.writerow(['Export Timestamp', self._get_export_timestamp(config)])
                writer.writerow([])

            if include_grain_size_stats and config.get('percentiles', True):
                writer.writerow(['Grain Size Percentiles'])
                writer.writerow(['Percentile', 'Size (mm)'])
                for _key, pname, percentile_num in self._selected_percentile_specs(config):
                    value = self._percentile_value(dataset, percentile_num)
                    if value is not None:
                        writer.writerow([pname, f"{value:.4f}"])
                writer.writerow([])

            if include_grain_size_stats and config.get('gradation', True):
                writer.writerow(['Gradation Parameters'])
                writer.writerow(['Parameter', 'Value'])
                cu = dataset.get_uniformity_coefficient() if hasattr(dataset, 'get_uniformity_coefficient') else None
                cc = dataset.get_coefficient_of_curvature() if hasattr(dataset, 'get_coefficient_of_curvature') else None
                if cu is not None:
                    writer.writerow(['Uniformity Coefficient (Cu)', f"{cu:.2f}"])
                if cc is not None:
                    writer.writerow(['Coefficient of Curvature (Cc)', f"{cc:.2f}"])
                writer.writerow([])

            if config.get('statistics', True) and filtered_results and stats_values and selected_stats:
                writer.writerow(['K-Value Statistics'])
                writer.writerow(['Statistic'] + [unit_label for _, unit_label, _, _, _, _ in unit_specs])
                for _, label, _, value_key in selected_stats:
                    if value_key == 'valid_count':
                        row = [label]
                        if unit_specs:
                            row.append(str(stats_values.get('valid_count', '')))
                            row.extend([''] * (len(unit_specs) - 1))
                        writer.writerow(row)
                        continue

                    stat_value = stats_values.get(value_key)
                    if stat_value is None:
                        continue

                    row = [label]
                    row.extend(self._format_converted_value(stat_value, unit_spec) for unit_spec in unit_specs)
                    writer.writerow(row)

    # ==================== EXCEL EXPORT ====================

    def _export_excel_single(self, name: str, dataset: GrainSizeData,
                            results: List[KCalculationResult], config: Dict):
        """Export single dataset to Excel workbook with multiple sheets"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        output_dir = self._category_output_dir(config, 'workbooks')
        template = config['filename_template']

        filename = self._format_filename(template, name, '.xlsx')
        filepath = os.path.join(output_dir, filename)

        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Sheet 1: Summary
        ws_summary = wb.create_sheet('Summary')
        self._write_excel_summary(ws_summary, name, dataset, results, config)

        # Sheet 2: Grain Size Data
        if config.get('grain_distribution', True):
            ws_grain = wb.create_sheet('Grain_Size_Data')
            self._write_excel_grain_size(ws_grain, dataset)

        # Sheet 3: Percentiles
        if config.get('percentiles', True):
            ws_percentiles = wb.create_sheet('Percentiles')
            self._write_excel_percentiles(ws_percentiles, dataset, config)

        # Sheet 4: K-Values
        if config.get('k_values', True) and results:
            ws_k = wb.create_sheet('K_Values')
            self._write_excel_k_values(ws_k, results, config)

        # Sheet 5: Statistics
        if config.get('statistics', True) and results:
            ws_stats = wb.create_sheet('Statistics')
            self._write_excel_statistics(ws_stats, results, config)

        wb.save(filepath)
        self.exported_files.append(filepath)

    def _export_excel_combined(self, datasets: List[tuple], config: Dict):
        """Export all datasets to single Excel workbook"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        output_dir = self._category_output_dir(config, 'workbooks')
        template = config['filename_template']

        filename = self._format_filename(template, 'combined_all_datasets', '.xlsx')
        filepath = os.path.join(output_dir, filename)

        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create a summary sheet comparing all datasets
        ws_summary = wb.create_sheet('All_Datasets_Summary')
        self._write_excel_combined_summary(ws_summary, datasets, config)

        # Create individual sheets for each dataset
        for name, dataset, results in datasets:
            # Sanitize sheet name (max 31 chars, no special chars)
            sheet_name = name[:31]
            for char in ['\\', '/', '*', '[', ']', ':', '?']:
                sheet_name = sheet_name.replace(char, '_')

            ws = wb.create_sheet(sheet_name)
            self._write_excel_dataset_combined(ws, name, dataset, results, config)

        wb.save(filepath)
        self.exported_files.append(filepath)

    def _export_excel_method_organized(self, datasets: List[tuple], config: Dict):
        """Export Excel workbook organized by K-calculation methods"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        output_dir = self._category_output_dir(config, 'workbooks')
        template = config['filename_template']

        filename = self._format_filename(template, 'method_comparison', '.xlsx')
        filepath = os.path.join(output_dir, filename)

        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Get all unique methods
        all_methods = set()
        for _, _, results in datasets:
            if results:
                for result in results:
                    all_methods.add(result.method_name)

        # Create a sheet for each method
        for method in sorted(all_methods):
            # Sanitize sheet name
            sheet_name = method[:31]
            ws = wb.create_sheet(sheet_name)
            self._write_excel_method_comparison(ws, method, datasets)

        wb.save(filepath)
        self.exported_files.append(filepath)

    def _write_excel_summary(self, ws, name: str, dataset: GrainSizeData,
                            results: List[KCalculationResult], config: Dict):
        """Write summary sheet to Excel"""
        from openpyxl.styles import Font, Alignment, PatternFill

        # Title
        ws['A1'] = 'Grain Size Analysis Summary'
        ws['A1'].font = Font(size=14, bold=True)

        row = 3

        # Sample Information
        ws[f'A{row}'] = 'Sample Name:'
        ws[f'B{row}'] = name
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = 'Temperature:'
        ws[f'B{row}'] = f"{dataset.temperature}°C"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = 'Porosity:'
        ws[f'B{row}'] = self._effective_porosity(dataset)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = 'Porosity Source:'
        ws[f'B{row}'] = self._porosity_source_label(dataset)
        ws[f'A{row}'].font = Font(bold=True)
        row += 2

        # Key grain sizes
        ws[f'A{row}'] = 'Key Grain Sizes'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        percentiles = {
            'D10': dataset.get_d10() if hasattr(dataset, 'get_d10') else None,
            'D50': dataset.get_d50() if hasattr(dataset, 'get_d50') else None,
            'D60': dataset.get_d60() if hasattr(dataset, 'get_d60') else None,
        }

        for pname, value in percentiles.items():
            if value is not None:
                ws[f'A{row}'] = f"{pname}:"
                ws[f'B{row}'] = f"{value:.4f} mm"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

        row += 1

        # Gradation
        cu = dataset.get_uniformity_coefficient() if hasattr(dataset, 'get_uniformity_coefficient') else None
        cc = dataset.get_coefficient_of_curvature() if hasattr(dataset, 'get_coefficient_of_curvature') else None

        if cu is not None:
            ws[f'A{row}'] = 'Uniformity Coefficient (Cu):'
            ws[f'B{row}'] = f"{cu:.2f}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        if cc is not None:
            ws[f'A{row}'] = 'Coefficient of Curvature (Cc):'
            ws[f'B{row}'] = f"{cc:.2f}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Classification
        if config.get('classification', True):
            row += 1
            _cls = dataset.classify(scheme=self._scheme)
            ws[f'A{row}'] = 'Soil Classification:'
            ws[f'B{row}'] = _cls.label
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = 'Standard:'
            ws[f'B{row}'] = _cls.scheme.name
            row += 1
            ws[f'A{row}'] = 'Clay %:'
            ws[f'B{row}'] = _cls.fractions.clay_pct
            row += 1
            ws[f'A{row}'] = 'Silt %:'
            ws[f'B{row}'] = _cls.fractions.silt_pct
            row += 1
            ws[f'A{row}'] = 'Sand %:'
            ws[f'B{row}'] = _cls.fractions.sand_pct
            row += 1
            ws[f'A{row}'] = 'Gravel %:'
            ws[f'B{row}'] = _cls.fractions.gravel_pct

        # K-value statistics
        if results:
            valid_k_values = [r.k_value for r in results if r.k_value is not None and r.k_value > 0]

            if valid_k_values:
                import statistics

                row += 2
                ws[f'A{row}'] = 'K-Value Statistics'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1

                mean_k = statistics.mean(valid_k_values)
                median_k = statistics.median(valid_k_values)

                ws[f'A{row}'] = 'Mean K:'
                ws[f'B{row}'] = f"{mean_k:.3e} m/s"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

                ws[f'A{row}'] = 'Median K:'
                ws[f'B{row}'] = f"{median_k:.3e} m/s"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

        # Auto-size columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _write_excel_grain_size(self, ws, dataset: GrainSizeData):
        """Write grain size data sheet"""
        from openpyxl.styles import Font

        # Header
        ws['A1'] = 'Particle Size (mm)'
        ws['B1'] = 'Percent Passing (%)'
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)

        # Data
        for i, (size, percent) in enumerate(zip(dataset.particle_sizes, dataset.percent_passing), start=2):
            ws[f'A{i}'] = size
            ws[f'B{i}'] = percent

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20

    def _write_excel_percentiles(self, ws, dataset: GrainSizeData, config: Dict):
        """Write percentiles sheet"""
        from openpyxl.styles import Font

        # Header
        ws['A1'] = 'Percentile'
        ws['B1'] = 'Size (mm)'
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)

        row = 2
        for _key, name, percentile_num in self._selected_percentile_specs(config):
            value = self._percentile_value(dataset, percentile_num)
            if value is not None:
                ws[f'A{row}'] = name
                ws[f'B{row}'] = value
                row += 1

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15

    def _write_excel_k_values(self, ws, results: List[KCalculationResult], config: Dict):
        """Write K-values sheet"""
        from openpyxl.styles import Font

        # Header
        unit_specs = self._get_enabled_unit_specs(config)
        filtered_results = self._filter_results(results, config)
        headers = ['Method']
        headers.extend(unit_label for _, unit_label, _, _, _, _ in unit_specs)
        headers.append('Status')
        if config.get('validation', False):
            headers.append('Status Message')
        if config.get('formulas', False):
            headers.append('Formula')
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # Data
        row = 2
        for result in filtered_results:
            col = 1
            ws.cell(row=row, column=col).value = result.method_name
            col += 1
            for _unit_key, _unit_label, _suffix, _json_key, multiplier, _fmt in unit_specs:
                ws.cell(row=row, column=col).value = result.k_value * multiplier
                col += 1
            ws.cell(row=row, column=col).value = result.status.value if hasattr(result.status, 'value') else str(result.status)
            col += 1
            if config.get('validation', False):
                ws.cell(row=row, column=col).value = result.status_message
                col += 1
            if config.get('formulas', False):
                ws.cell(row=row, column=col).value = result.formula_used
            row += 1

        # Auto-size
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _write_excel_statistics(self, ws, results: List[KCalculationResult]):
        """Write statistics sheet"""
        from openpyxl.styles import Font
        import statistics

        valid_k_values = [r.k_value for r in results if r.k_value is not None and r.k_value > 0]

        if not valid_k_values:
            return

        # Header
        ws['A1'] = 'Statistic'
        ws['B1'] = 'Value (m/s)'
        ws['C1'] = 'Value (cm/s)'
        ws['D1'] = 'Value (m/d)'

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}1'].font = Font(bold=True)

        # Calculate statistics
        mean_k = statistics.mean(valid_k_values)
        median_k = statistics.median(valid_k_values)
        min_k = min(valid_k_values)
        max_k = max(valid_k_values)
        stdev_k = statistics.stdev(valid_k_values) if len(valid_k_values) > 1 else 0

        stats_data = [
            ('Mean', mean_k),
            ('Median', median_k),
            ('Std Dev', stdev_k),
            ('Min', min_k),
            ('Max', max_k),
        ]

        for row, (name, value) in enumerate(stats_data, start=2):
            ws[f'A{row}'] = name
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'C{row}'] = value * 100
            ws[f'D{row}'] = value * 86400

        # Auto-size
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 15

    def _write_excel_combined_summary(self, ws, datasets: List[tuple], config: Dict):
        """Write combined summary comparing all datasets"""
        from openpyxl.styles import Font

        # Header
        ws['A1'] = 'Sample'
        headers = ['Sample']
        col = 2

        # Get all method names
        all_methods = set()
        for _, _, results in datasets:
            if results:
                for result in results:
                    all_methods.add(result.method_name)

        for method in sorted(all_methods):
            ws.cell(row=1, column=col).value = method
            ws.cell(row=1, column=col).font = Font(bold=True)
            col += 1

        # Add statistical columns
        ws.cell(row=1, column=col).value = 'Mean'
        ws.cell(row=1, column=col).font = Font(bold=True)
        col += 1
        ws.cell(row=1, column=col).value = 'Median'
        ws.cell(row=1, column=col).font = Font(bold=True)

        # Data rows
        for row, (name, dataset, results) in enumerate(datasets, start=2):
            ws.cell(row=row, column=1).value = name

            # Create method -> k_value mapping
            method_values = {}
            if results:
                for result in results:
                    if result.k_value is not None:
                        method_values[result.method_name] = result.k_value

            # Fill in K-values
            col = 2
            for method in sorted(all_methods):
                if method in method_values:
                    ws.cell(row=row, column=col).value = method_values[method]
                col += 1

            # Calculate and fill statistics
            if method_values:
                import statistics
                values = list(method_values.values())
                ws.cell(row=row, column=col).value = statistics.mean(values)
                col += 1
                ws.cell(row=row, column=col).value = statistics.median(values)

    def _write_excel_dataset_combined(self, ws, name: str, dataset: GrainSizeData,
                                     results: List[KCalculationResult], config: Dict):
        """Write all data for a single dataset on one sheet"""
        from openpyxl.styles import Font

        row = 1

        # Title
        ws[f'A{row}'] = name
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2

        # Grain sizes
        ws[f'A{row}'] = 'Key Grain Sizes'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        percentiles = {
            'D10': dataset.get_d10() if hasattr(dataset, 'get_d10') else None,
            'D50': dataset.get_d50() if hasattr(dataset, 'get_d50') else None,
            'D60': dataset.get_d60() if hasattr(dataset, 'get_d60') else None,
        }

        for pname, value in percentiles.items():
            if value is not None:
                ws[f'A{row}'] = pname
                ws[f'B{row}'] = value
                row += 1

        row += 1

        # K-values
        ws[f'A{row}'] = 'K-Values'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = 'Method'
        ws[f'B{row}'] = 'K (m/s)'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        if results:
            for result in results:
                if result.k_value is not None:
                    ws[f'A{row}'] = result.method_name
                    ws[f'B{row}'] = result.k_value
                    row += 1

    def _write_excel_method_comparison(self, ws, method: str, datasets: List[tuple]):
        """Write comparison sheet for a specific method across all datasets"""
        from openpyxl.styles import Font

        # Header
        ws['A1'] = 'Sample'
        ws['B1'] = f'{method} K (m/s)'
        ws['C1'] = 'K (cm/s)'
        ws['D1'] = 'K (m/d)'
        ws['E1'] = 'Status'

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}1'].font = Font(bold=True)

        # Data
        row = 2
        for name, dataset, results in datasets:
            if results:
                for result in results:
                    if result.method_name == method and result.k_value is not None:
                        ws[f'A{row}'] = name
                        ws[f'B{row}'] = result.k_value
                        ws[f'C{row}'] = result.k_value * 100
                        ws[f'D{row}'] = result.k_value * 86400
                        ws[f'E{row}'] = result.status.value if hasattr(result.status, 'value') else str(result.status)
                        row += 1
                        break

        # Auto-size
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18

    # ==================== JSON EXPORT ====================

    def _export_json(self, name: str, dataset: GrainSizeData,
                    results: List[KCalculationResult], config: Dict):
        """Export dataset to JSON format"""
        output_dir = self._category_output_dir(config, 'data', 'json')
        template = config['filename_template']

        filename = self._format_filename(template, name, '.json')
        filepath = os.path.join(output_dir, filename)

        # Build JSON structure
        data = {
            'sample_name': name,
            'metadata': {
                'temperature': dataset.temperature,
                'porosity': self._effective_porosity(dataset),
                'porosity_source': self._porosity_source_label(dataset),
                'file_path': dataset.file_path if hasattr(dataset, 'file_path') else None,
            }
        }

        # Grain size distribution
        if config.get('grain_distribution', True):
            data['grain_size_distribution'] = {
                'particle_sizes_mm': dataset.particle_sizes,
                'percent_passing': dataset.percent_passing,
            }

        # Percentiles
        if config.get('percentiles', True):
            data['percentiles'] = {
                'D10': dataset.get_d10() if hasattr(dataset, 'get_d10') else None,
                'D20': dataset.get_d20() if hasattr(dataset, 'get_d20') else None,
                'D30': dataset.get_d30() if hasattr(dataset, 'get_d30') else None,
                'D50': dataset.get_d50() if hasattr(dataset, 'get_d50') else None,
                'D60': dataset.get_d60() if hasattr(dataset, 'get_d60') else None,
            }

        # Gradation
        if config.get('gradation', True):
            data['gradation'] = {
                'uniformity_coefficient': dataset.get_uniformity_coefficient() if hasattr(dataset, 'get_uniformity_coefficient') else None,
                'coefficient_of_curvature': dataset.get_coefficient_of_curvature() if hasattr(dataset, 'get_coefficient_of_curvature') else None,
            }

        # Classification
        if config.get('classification', True):
            _cls = dataset.classify(scheme=self._scheme)
            data['classification'] = {
                'label':       _cls.label,
                'scheme_name': _cls.scheme.name,
                'clay_pct':    _cls.fractions.clay_pct,
                'silt_pct':    _cls.fractions.silt_pct,
                'sand_pct':    _cls.fractions.sand_pct,
                'gravel_pct':  _cls.fractions.gravel_pct,
                'cobble_pct':  _cls.fractions.cobble_pct,
            }

        # K-values
        if config.get('k_values', True) and results:
            data['k_values'] = []
            for result in results:
                if result.k_value is not None:
                    k_data = {
                        'method': result.method_name,
                        'k_m_s': result.k_value,
                        'k_cm_s': result.k_value * 100,
                        'k_m_d': result.k_value * 86400,
                        'status': result.status.value if hasattr(result.status, 'value') else str(result.status),
                        'status_message': result.status_message,
                        'grain_size_used': result.grain_size_used,
                    }

                    if config.get('formulas', False):
                        k_data['formula'] = result.formula_used

                    data['k_values'].append(k_data)

        # Statistics
        if config.get('statistics', True) and results:
            valid_k_values = [r.k_value for r in results if r.k_value is not None and r.k_value > 0]

            if valid_k_values:
                import statistics

                data['statistics'] = {
                    'mean_k_m_s': statistics.mean(valid_k_values),
                    'median_k_m_s': statistics.median(valid_k_values),
                    'min_k_m_s': min(valid_k_values),
                    'max_k_m_s': max(valid_k_values),
                    'stdev_k_m_s': statistics.stdev(valid_k_values) if len(valid_k_values) > 1 else 0,
                    'count': len(valid_k_values),
                }

        # Write JSON file
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)

        self.exported_files.append(filepath)

    def _write_excel_summary(self, ws, name: str, dataset: GrainSizeData,
                            results: List[KCalculationResult], config: Dict):
        """Write summary sheet to Excel."""
        from openpyxl.styles import Font

        ws['A1'] = 'Grain Size Analysis Summary'
        ws['A1'].font = Font(size=14, bold=True)

        row = 3
        include_sample_info = self._metadata_enabled(config, 'sample_info')
        include_environmental = self._metadata_enabled(config, 'environmental')
        include_timestamp = self._metadata_enabled(config, 'export_timestamp')
        include_grain_size_stats = config.get('include_grain_size_stats', True)
        unit_specs = self._get_enabled_unit_specs(config)
        primary_unit = unit_specs[0] if unit_specs else None

        if include_sample_info:
            ws[f'A{row}'] = 'Sample Name:'
            ws[f'B{row}'] = name
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        if include_environmental:
            ws[f'A{row}'] = 'Temperature:'
            ws[f'B{row}'] = f"{dataset.temperature} C"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = 'Porosity:'
            ws[f'B{row}'] = self._effective_porosity(dataset)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = 'Porosity Source:'
            ws[f'B{row}'] = self._porosity_source_label(dataset)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        if include_timestamp:
            ws[f'A{row}'] = 'Exported At:'
            ws[f'B{row}'] = self._get_export_timestamp(config)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        if include_sample_info or include_environmental or include_timestamp:
            row += 1

        if include_grain_size_stats and config.get('percentiles', True):
            ws[f'A{row}'] = 'Key Grain Sizes'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            for _key, percentile_name, percentile_num in self._selected_percentile_specs(config):
                value = self._percentile_value(dataset, percentile_num)
                if value is not None:
                    ws[f'A{row}'] = f'{percentile_name}:'
                    ws[f'B{row}'] = f'{value:.4f} mm'
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1

            row += 1

        if include_grain_size_stats and config.get('gradation', True):
            cu = dataset.get_uniformity_coefficient() if hasattr(dataset, 'get_uniformity_coefficient') else None
            cc = dataset.get_coefficient_of_curvature() if hasattr(dataset, 'get_coefficient_of_curvature') else None

            if cu is not None:
                ws[f'A{row}'] = 'Uniformity Coefficient (Cu):'
                ws[f'B{row}'] = f'{cu:.2f}'
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

            if cc is not None:
                ws[f'A{row}'] = 'Coefficient of Curvature (Cc):'
                ws[f'B{row}'] = f'{cc:.2f}'
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

        if config.get('classification', True):
            row += 1
            _cls = dataset.classify(scheme=self._scheme)
            ws[f'A{row}'] = 'Soil Classification:'
            ws[f'B{row}'] = _cls.label
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = 'Standard:'
            ws[f'B{row}'] = _cls.scheme.name
            row += 1
            ws[f'A{row}'] = 'Clay %:'
            ws[f'B{row}'] = _cls.fractions.clay_pct
            row += 1
            ws[f'A{row}'] = 'Silt %:'
            ws[f'B{row}'] = _cls.fractions.silt_pct
            row += 1
            ws[f'A{row}'] = 'Sand %:'
            ws[f'B{row}'] = _cls.fractions.sand_pct
            row += 1
            ws[f'A{row}'] = 'Gravel %:'
            ws[f'B{row}'] = _cls.fractions.gravel_pct

        stats_values = self._calculate_statistics(results, config) if config.get('statistics', True) else {}
        selected_stats = self._get_selected_stat_specs(config) if config.get('statistics', True) else []
        if stats_values and selected_stats:
            row += 2
            ws[f'A{row}'] = 'K-Value Statistics'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            for _, label, _, value_key in selected_stats:
                if value_key == 'valid_count':
                    ws[f'A{row}'] = f'{label}:'
                    ws[f'B{row}'] = stats_values['valid_count']
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1
                    continue

                if primary_unit is None:
                    continue

                stat_value = stats_values.get(value_key)
                if stat_value is None:
                    continue

                _, _, unit_suffix, _, multiplier, fmt = primary_unit
                ws[f'A{row}'] = f'{label} ({unit_suffix}):'
                ws[f'B{row}'] = format(stat_value * multiplier, fmt)
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _write_excel_statistics(self, ws, results: List[KCalculationResult], config: Dict):
        """Write statistics sheet."""
        from openpyxl.styles import Font

        unit_specs = self._get_enabled_unit_specs(config)
        stats_values = self._calculate_statistics(results, config)
        selected_stats = self._get_selected_stat_specs(config) if config.get('statistics', True) else []

        if not stats_values or not selected_stats:
            return

        headers = ['Statistic']
        if unit_specs:
            headers.extend(unit_label for _, unit_label, _, _, _, _ in unit_specs)
        else:
            headers.append('Value')

        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col).value = header
            ws.cell(row=1, column=col).font = Font(bold=True)

        row = 2
        for _, label, _, value_key in selected_stats:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = Font(bold=True)

            if value_key == 'valid_count':
                ws.cell(row=row, column=2).value = stats_values['valid_count']
                row += 1
                continue

            stat_value = stats_values.get(value_key)
            if stat_value is None:
                row += 1
                continue

            if unit_specs:
                for col, unit_spec in enumerate(unit_specs, start=2):
                    _, _, _, _, multiplier, _ = unit_spec
                    ws.cell(row=row, column=col).value = stat_value * multiplier
            else:
                ws.cell(row=row, column=2).value = stat_value

            row += 1

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _write_excel_combined_summary(self, ws, datasets: List[tuple], config: Dict):
        """Write combined summary comparing all datasets."""
        from openpyxl.styles import Font

        ws['A1'] = 'Sample'
        ws['A1'].font = Font(bold=True)
        col = 2
        primary_unit = self._get_primary_unit_spec(config)
        k_values_enabled = config.get('k_values', True)
        selected_stats = self._get_selected_stat_specs(config) if config.get('statistics', True) else []

        all_methods = set()
        if k_values_enabled:
            for _, _, results in datasets:
                for result in self._filter_results(results, config):
                    if result.k_value is not None:
                        all_methods.add(result.method_name)

        ordered_methods = self._ordered_method_names(all_methods)
        for method in ordered_methods:
            ws.cell(row=1, column=col).value = method
            ws.cell(row=1, column=col).font = Font(bold=True)
            col += 1

        for _, label, _, value_key in selected_stats:
            header = label if value_key == 'valid_count' else f'{label} ({primary_unit[2]})'
            ws.cell(row=1, column=col).value = header
            ws.cell(row=1, column=col).font = Font(bold=True)
            col += 1

        for row, (name, dataset, results) in enumerate(datasets, start=2):
            ws.cell(row=row, column=1).value = name

            method_values = {}
            if k_values_enabled:
                for result in self._filter_results(results, config):
                    if result.k_value is not None:
                        method_values[result.method_name] = result.k_value

            col = 2
            for method in ordered_methods:
                if method in method_values:
                    ws.cell(row=row, column=col).value = method_values[method]
                col += 1

            stats_values = self._calculate_statistics(results, config) if selected_stats else {}
            for _, _, _, value_key in selected_stats:
                if value_key == 'valid_count':
                    ws.cell(row=row, column=col).value = stats_values.get('valid_count')
                else:
                    stat_value = stats_values.get(value_key)
                    if stat_value is not None:
                        ws.cell(row=row, column=col).value = stat_value * primary_unit[4]
                col += 1

    def _export_json(self, name: str, dataset: GrainSizeData,
                    results: List[KCalculationResult], config: Dict):
        """Export dataset to JSON format."""
        output_dir = self._category_output_dir(config, 'data', 'json')
        template = config['filename_template']
        unit_specs = self._get_enabled_unit_specs(config)
        filtered_results = self._filter_results(results, config)
        selected_stats = self._get_selected_stat_specs(config) if config.get('statistics', True) else []
        stats_values = self._calculate_statistics(results, config) if selected_stats else {}

        filename = self._format_filename(template, name, '.json')
        filepath = os.path.join(output_dir, filename)

        data = {}
        if self._metadata_enabled(config, 'sample_info'):
            data['sample_name'] = name

        metadata = {}
        if self._metadata_enabled(config, 'sample_info') and hasattr(dataset, 'file_path') and dataset.file_path:
            metadata['file_path'] = dataset.file_path
        if self._metadata_enabled(config, 'environmental'):
            metadata['temperature'] = dataset.temperature
            metadata['porosity'] = self._effective_porosity(dataset)
            metadata['porosity_source'] = self._porosity_source_label(dataset)
        if self._metadata_enabled(config, 'export_timestamp'):
            metadata['exported_at'] = self._get_export_timestamp(config)
        if metadata:
            data['metadata'] = metadata

        if config.get('grain_distribution', True):
            data['grain_size_distribution'] = {
                'particle_sizes_mm': dataset.particle_sizes,
                'percent_passing': dataset.percent_passing,
            }

        if config.get('percentiles', True):
            data['percentiles'] = {
                label: self._percentile_value(dataset, percentile_num)
                for _key, label, percentile_num in self._selected_percentile_specs(config)
            }

        if config.get('gradation', True):
            data['gradation'] = {
                'uniformity_coefficient': dataset.get_uniformity_coefficient() if hasattr(dataset, 'get_uniformity_coefficient') else None,
                'coefficient_of_curvature': dataset.get_coefficient_of_curvature() if hasattr(dataset, 'get_coefficient_of_curvature') else None,
            }

        if config.get('classification', True):
            _cls = dataset.classify(scheme=self._scheme)
            data['classification'] = {
                'label':       _cls.label,
                'scheme_name': _cls.scheme.name,
                'clay_pct':    _cls.fractions.clay_pct,
                'silt_pct':    _cls.fractions.silt_pct,
                'sand_pct':    _cls.fractions.sand_pct,
                'gravel_pct':  _cls.fractions.gravel_pct,
                'cobble_pct':  _cls.fractions.cobble_pct,
            }

        if config.get('k_values', True) and filtered_results:
            data['k_values'] = []
            for result in filtered_results:
                k_data = {
                    'method': result.method_name,
                    'status': result.status.value if hasattr(result.status, 'value') else str(result.status),
                    'grain_size_used': result.grain_size_used,
                }

                if config.get('validation', False):
                    k_data['status_message'] = result.status_message
                if config.get('formulas', False):
                    k_data['formula'] = result.formula_used

                for unit_key, _, _, json_unit_key, multiplier, _ in unit_specs:
                    k_data[json_unit_key] = result.k_value * multiplier

                data['k_values'].append(k_data)

        if config.get('statistics', True) and stats_values and selected_stats:
            statistics_data = {}
            for _, _, _, value_key in selected_stats:
                if value_key == 'valid_count':
                    statistics_data['valid_count'] = stats_values['valid_count']
                    continue

                stat_value = stats_values.get(value_key)
                if stat_value is None:
                    continue

                for unit_key, _, _, _, multiplier, _ in unit_specs:
                    statistics_data[f'{value_key}_{unit_key}'] = stat_value * multiplier

            if statistics_data:
                data['statistics'] = statistics_data

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, default=str)

        self.exported_files.append(filepath)
