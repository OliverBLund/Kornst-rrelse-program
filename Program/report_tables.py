"""Report-table extraction, assessment, and Excel appendix rendering."""

from __future__ import annotations

from dataclasses import dataclass, field
from html.parser import HTMLParser
import io
import re
from typing import Optional


EXCEL_ROW_THRESHOLD = 50
EXCEL_CELL_THRESHOLD = 400


@dataclass(slots=True)
class ReportTableCell:
    text: str
    is_header: bool = False
    colspan: int = 1
    rowspan: int = 1


@dataclass(slots=True)
class ReportTable:
    table_id: str
    title: str
    classes: tuple[str, ...] = ()
    rows: list[list[ReportTableCell]] = field(default_factory=list)

    @property
    def column_count(self) -> int:
        return max(
            (sum(max(1, cell.colspan) for cell in row) for row in self.rows),
            default=0,
        )

    @property
    def header_row_count(self) -> int:
        count = 0
        for row in self.rows:
            if row and all(cell.is_header for cell in row):
                count += 1
            else:
                break
        return count

    @property
    def body_row_count(self) -> int:
        return max(0, len(self.rows) - self.header_row_count)

    @property
    def data_cell_count(self) -> int:
        return self.body_row_count * self.column_count

    @property
    def excel_recommended(self) -> bool:
        return (
            self.body_row_count > EXCEL_ROW_THRESHOLD
            or self.data_cell_count > EXCEL_CELL_THRESHOLD
        )


@dataclass(slots=True)
class ReportTableAnalysis:
    tables: list[ReportTable]

    @property
    def large_tables(self) -> list[ReportTable]:
        return [table for table in self.tables if table.excel_recommended]

    @property
    def excel_recommended(self) -> bool:
        return bool(self.large_tables)

    @property
    def recommendation_text(self) -> str:
        large = self.large_tables
        if not large:
            return ""
        names = ", ".join(table.title or table.table_id for table in large[:3])
        suffix = "" if len(large) <= 3 else f" and {len(large) - 3} more"
        return (
            f"{len(large)} large table{'s' if len(large) != 1 else ''} "
            f"would be easier to work with in Excel: {names}{suffix}."
        )


def _clean_text(value: str) -> str:
    lines = [
        re.sub(r"[ \t\r\f\v]+", " ", line).strip()
        for line in value.split("\n")
    ]
    return "\n".join(line for line in lines if line)


def _slug(value: str, fallback: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or fallback


class _ReportTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[ReportTable] = []
        self._heading_tag: Optional[str] = None
        self._heading_parts: list[str] = []
        self._latest_heading = ""
        self._table: Optional[ReportTable] = None
        self._row: Optional[list[ReportTableCell]] = None
        self._cell: Optional[ReportTableCell] = None
        self._cell_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        attrs_dict = {key: value or "" for key, value in attrs}
        if tag in {"h1", "h2", "h3", "h4"}:
            self._heading_tag = tag
            self._heading_parts = []
            return
        if tag == "table":
            index = len(self.tables) + 1
            title = self._latest_heading or f"Table {index}"
            table_id = attrs_dict.get("data-report-table") or _slug(
                title, f"table-{index}"
            )
            self._table = ReportTable(
                table_id=table_id,
                title=title,
                classes=tuple(attrs_dict.get("class", "").split()),
            )
            return
        if tag == "tr" and self._table is not None:
            self._row = []
            return
        if tag in {"th", "td"} and self._row is not None:
            self._cell = ReportTableCell(
                text="",
                is_header=tag == "th",
                colspan=max(1, int(attrs_dict.get("colspan", "1") or 1)),
                rowspan=max(1, int(attrs_dict.get("rowspan", "1") or 1)),
            )
            self._cell_parts = []
            return
        if tag == "br" and self._cell is not None:
            self._cell_parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self._heading_tag is not None:
            self._heading_parts.append(data)
        if self._cell is not None:
            self._cell_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == self._heading_tag:
            self._latest_heading = _clean_text("".join(self._heading_parts))
            self._heading_tag = None
            self._heading_parts = []
            return
        if tag in {"th", "td"} and self._cell is not None:
            self._cell.text = _clean_text("".join(self._cell_parts))
            if self._row is not None:
                self._row.append(self._cell)
            self._cell = None
            self._cell_parts = []
            return
        if tag == "tr" and self._row is not None:
            if self._table is not None:
                self._table.rows.append(self._row)
            self._row = None
            return
        if tag == "table" and self._table is not None:
            self.tables.append(self._table)
            self._table = None


def extract_report_tables(html_text: str) -> list[ReportTable]:
    parser = _ReportTableParser()
    parser.feed(html_text)
    parser.close()
    return parser.tables


def analyze_report_tables(html_text: str) -> ReportTableAnalysis:
    return ReportTableAnalysis(extract_report_tables(html_text))


def _safe_sheet_name(value: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]+", " ", value).strip() or "Table"
    base = re.sub(r"\s+", " ", base)[:31]
    candidate = base
    suffix = 2
    while candidate.lower() in used:
        trailer = f" {suffix}"
        candidate = f"{base[:31 - len(trailer)]}{trailer}"
        suffix += 1
    used.add(candidate.lower())
    return candidate


def generate_excel_appendix(
    html_text: str,
    title: str = "Report data appendix",
    accent_color: str = "#2c3e50",
) -> bytes:
    """Create one professionally formatted worksheet per report table."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError("openpyxl is required for Excel appendices.") from exc

    tables = extract_report_tables(html_text)
    if not tables:
        raise ValueError("The report contains no tables to export.")

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names: set[str] = set()

    accent = re.sub(r"[^0-9A-Fa-f]", "", str(accent_color or ""))
    dark = accent.upper() if len(accent) == 6 else "2C3E50"
    red, green, blue = (int(dark[index:index + 2], 16) for index in (0, 2, 4))
    pale = "".join(
        f"{round(channel + (255 - channel) * 0.90):02X}"
        for channel in (red, green, blue)
    )
    rule = Side(style="thin", color="7D8992")
    strong_rule = Side(style="medium", color=dark)

    for table_index, table in enumerate(tables, start=1):
        sheet = workbook.create_sheet(
            _safe_sheet_name(table.title or f"Table {table_index}", used_names)
        )
        sheet.sheet_view.showGridLines = False
        occupied: set[tuple[int, int]] = set()

        for row_index, row in enumerate(table.rows, start=1):
            column_index = 1
            for source_cell in row:
                while (row_index, column_index) in occupied:
                    column_index += 1

                cell = sheet.cell(row=row_index, column=column_index)
                cell.value = source_cell.text
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
                if source_cell.is_header:
                    cell.font = Font(bold=True, color=dark)
                    cell.fill = PatternFill("solid", fgColor=pale)
                    cell.border = Border(bottom=strong_rule)

                end_row = row_index + source_cell.rowspan - 1
                end_col = column_index + source_cell.colspan - 1
                if end_row > row_index or end_col > column_index:
                    sheet.merge_cells(
                        start_row=row_index,
                        start_column=column_index,
                        end_row=end_row,
                        end_column=end_col,
                    )
                for occupied_row in range(row_index, end_row + 1):
                    for occupied_col in range(column_index, end_col + 1):
                        occupied.add((occupied_row, occupied_col))
                column_index = end_col + 1

        if table.rows:
            last_row = len(table.rows)
            for column_index in range(1, table.column_count + 1):
                sheet.cell(row=last_row, column=column_index).border = Border(
                    bottom=strong_rule
                )

        for column_index in range(1, table.column_count + 1):
            values = [
                str(sheet.cell(row=row_index, column=column_index).value or "")
                for row_index in range(1, min(sheet.max_row, 80) + 1)
            ]
            width = min(max(max((len(line) for value in values for line in value.splitlines()), default=0) + 2, 10), 42)
            sheet.column_dimensions[get_column_letter(column_index)].width = width

        header_rows = table.header_row_count
        if header_rows:
            sheet.freeze_panes = sheet.cell(row=header_rows + 1, column=1)
            sheet.auto_filter.ref = (
                f"A{header_rows}:{get_column_letter(table.column_count)}{sheet.max_row}"
            )

        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5

    workbook.properties.title = title
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
