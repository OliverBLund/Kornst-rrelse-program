"""Typed table model shared by CSV preview/export and Excel export."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Sequence


@dataclass(frozen=True)
class ExportTable:
    """Rectangular export table with typed cell values.

    CSV has no real cell types, but keeping numbers as numbers until the final
    writer lets Excel exports receive numeric cells and keeps previews/export
    paths on the same schema.
    """

    name: str
    headers: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]

    @classmethod
    def from_rows(cls, name: str, headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> "ExportTable":
        header_tuple = tuple(str(header) for header in headers)
        row_tuples = tuple(tuple(row) for row in rows)
        return cls(name=name, headers=header_tuple, rows=row_tuples)

    def as_rows(self, max_data_rows: int | None = None) -> list[list[Any]]:
        body = self.rows if max_data_rows is None else self.rows[:max_data_rows]
        return [list(self.headers)] + [list(row) for row in body]


def write_csv_table(path: str | Path, table: ExportTable) -> None:
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerows(table.as_rows())


def write_excel_table(ws: Any, table: ExportTable) -> None:
    from openpyxl.styles import Font

    for col, header in enumerate(table.headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    for row_index, row in enumerate(table.rows, start=2):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index).value = value

    for col_index, header in enumerate(table.headers, start=1):
        column_letter = ws.cell(row=1, column=col_index).column_letter
        ws.column_dimensions[column_letter].width = min(max(len(str(header)) + 2, 12), 34)
